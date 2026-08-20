#!/usr/bin/env python3
"""
Empirical tests for the noise-meter data layer, run against real temporary
SQLite databases built from the MEAS118 reference SD-card files.

These are deliberately not import-smoke tests: every case drives real data
through the real code path and checks values, because the bugs this codebase
has actually shipped (a placeholder count off by one, a percentile reading the
wrong channel, a sync path dropping metadata) all compiled cleanly and read
correctly.

Replication is tested with two genuinely independent copies of the data layer,
each bound to its own database file, so a "peer" really is a separate Pi.

    python3 test_modules.py [path/to/MEAS118]

Exits non-zero on the first failure.
"""
import importlib
import json
import os
import shutil
import sys
import tempfile
import types

REPO = os.path.dirname(os.path.abspath(__file__))
MEAS_DEFAULT = os.path.join(REPO, 'MEAS118')
DATE_FOLDER = '260812'
SESSION_DATE = '2026-08-12'

# noise_app refuses to start when the flight tracker's auth module is missing,
# which it always is on a development Mac. Sections 4d and 8 import the real
# app, so the suite declares itself a development instance. Section 8 checks
# that a *subprocess* without this variable is still refused.
os.environ.setdefault('ALLOW_UNAUTHENTICATED', '1')

# The CSRF token the suite's test clients present. Seeded straight into the
# session rather than scraped out of a rendered page.
SUITE_CSRF = 'suite-csrf-token'
CSRF_HDR = {'X-CSRF-Token': SUITE_CSRF}

# Half-up rounding, reimplemented rather than imported: the point of the
# rounding checks in section 11 is that the app rounds .x5 away from zero, and
# importing the app's own helper to compute the expected value would assert
# nothing. Matches nor140_format.round_half_up by construction.
def _rhu(value, digits=1):
    import math as _m
    scale = 10 ** digits
    return _m.floor(value * scale + 0.5) / scale

_checks = 0


def check(cond, label, detail=''):
    global _checks
    _checks += 1
    if not cond:
        print(f'  FAIL  {label}')
        if detail:
            print(f'        {detail}')
        raise SystemExit(1)
    print(f'  ok    {label}' + (f'  ({detail})' if detail else ''))


def close(a, b, tol, label):
    check(a is not None and abs(a - b) <= tol, label, f'got {a!r}, expected {b} ±{tol}')


class Side:
    """One Pi: an independent import of the data layer bound to its own DB."""

    def __init__(self, db_path):
        self.db_path = db_path
        os.environ['NOISE_DB_PATH'] = db_path
        for name in ('noise_db', 'reports_db', 'assessments_db', 'sync_db', 'reports'):
            sys.modules.pop(name, None)
        self.db = importlib.import_module('noise_db')
        assert self.db.DB_PATH == db_path, self.db.DB_PATH
        self.sync = importlib.import_module('sync_db')
        self.assess = importlib.import_module('assessments_db')
        self.reports_db = importlib.import_module('reports_db')
        self.reports = importlib.import_module('reports')
        self.db.init_db()

    def sql(self, query, *params):
        conn = self.db.get_db()
        rows = conn.execute(query, params).fetchall()
        conn.close()
        return rows

    def exec(self, query, *params):
        conn = self.db.get_db()
        conn.execute(query, params)
        conn.commit()
        conn.close()

    def exec_nofk(self, query, *params):
        """Write with foreign keys OFF — how orphaned rows arise in the first place.
        noise_db.get_db() always enables them, so an orphan cannot be created
        through it; this reproduces the connection that caused the real ones."""
        import sqlite3
        conn = sqlite3.connect(self.db_path)
        conn.execute(query, params)
        conn.commit()
        conn.close()

    def has_session(self, date):
        return bool(self.sql('SELECT 1 FROM sessions WHERE date=?', date))

    def tombstones(self):
        return {r['date']: r['deleted_at'] for r in self.sql(
            'SELECT date, deleted_at FROM deleted_sessions')}

    def set_imported_at(self, date, ts):
        self.exec('UPDATE sessions SET imported_at=? WHERE date=?', ts, date)


REF_DIR = os.path.expanduser('~/Downloads/nortfr')

# Two things Nortfr emits that we cannot yet reproduce. Everything else must match
# exactly, so these are named narrowly rather than tolerated as a diff budget.
#   - the per-period marker column ('Pause', or the code 4), present only in the
#     2668-byte GLOB variant; its extra block is decoded only as far as the last
#     marked period, so the marked region cannot be rebuilt.
#   - Sensitivity, which varies per measurement (-26.9 / -27.0 / -27.4 seen) and
#     whose offset has not been located.
def _is_known_gap(sheet, row, col, gen, ref, ncols):
    if sheet == 'Summary' and gen is None and (ref == 'Pause' or ref == 4):
        return True                       # marker column
    if sheet == 'Setup' and row == 14:
        return True                       # Sensitivity
    return False


def check_references(meas_root, tmp):
    """Diff our export against every reference pair in ~/Downloads/nortfr."""
    import collections
    import glob as _glob
    import io
    import re as _re
    from openpyxl import load_workbook
    from nor140_exporter import build_global_xlsx, build_profile_xlsx
    from noise_parser import parse_files

    pairs = collections.defaultdict(dict)
    for p in _glob.glob(os.path.join(REF_DIR, '*.xlsx')):
        m = _re.search(r'_(\d{6})_(\d{4})_(GLOBAL|PROFILE)(-\d+)?\.xlsx$', os.path.basename(p))
        if m and not m.group(4):
            pairs[(m.group(1), int(m.group(2)))][m.group(3)] = p
    if not pairs:
        print('  SKIP — no Nortfr reference workbooks in ~/Downloads/nortfr')
        return

    def cells(obj):
        wb = load_workbook(io.BytesIO(obj) if isinstance(obj, bytes) else obj, data_only=True)
        out = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)] for ws in wb.worksheets}
        wb.close()
        return out

    by_date = collections.defaultdict(list)
    for d, rn in sorted(pairs):
        by_date[d].append(rn)

    for date_folder in sorted(by_date):
        root = os.path.join(meas_root, date_folder, 'PART0000')
        if not os.path.isdir(root):
            print(f'  SKIP {date_folder} — not in MEAS118')
            continue
        side = Side(os.path.join(tmp, f'ref-{date_folder}.db'))
        files = []
        for proj in sorted(os.listdir(root)):
            d = os.path.join(root, proj)
            if os.path.isdir(d):
                for fn in sorted(os.listdir(d)):
                    with open(os.path.join(d, fn), 'rb') as fh:
                        files.append((f'{date_folder}/PART0000/{proj}/{fn}', fh.read()))
        side.db.import_sessions(parse_files(files))
        iso = f'20{date_folder[0:2]}-{date_folder[2:4]}-{date_folder[4:6]}'

        for rn in by_date[date_folder]:
            # Pair on source_file: run_number is a sequential import index and
            # diverges from the PROJ folder number on dates with gaps.
            rows = side.sql('SELECT r.* FROM runs r JOIN sessions s ON r.session_id=s.id '
                            'WHERE s.date=? AND UPPER(r.source_file)=?', iso, 'PROJ%04d' % rn)
            check(bool(rows), f'{date_folder} run {rn}: present in the DB')
            run = dict(rows[0])
            run['session_date'] = iso
            for label, built in (('GLOBAL', build_global_xlsx(run, '6899108')),
                                 ('PROFILE', build_profile_xlsx(run, '6899108'))):
                if label not in pairs[(date_folder, rn)]:
                    continue
                gen, ref = cells(built), cells(pairs[(date_folder, rn)][label])
                check(list(gen) == list(ref),
                      f'{date_folder}/{rn} {label}: sheet names and order match')
                unexpected = []
                for name in ref:
                    g, r_ = gen.get(name, []), ref[name]
                    for i in range(max(len(g), len(r_))):
                        gr = g[i] if i < len(g) else []
                        rr = r_[i] if i < len(r_) else []
                        for j in range(max(len(gr), len(rr))):
                            gv = gr[j] if j < len(gr) else None
                            rv = rr[j] if j < len(rr) else None
                            if gv == rv:
                                continue
                            if isinstance(gv, float) and isinstance(rv, float) and abs(gv - rv) < 1e-9:
                                continue
                            if _is_known_gap(name, i + 1, j + 1, gv, rv, len(rr)):
                                continue
                            unexpected.append(f'{name} r{i+1}c{j+1}: {gv!r} vs {rv!r}')
                check(not unexpected,
                      f'{date_folder}/{rn} {label}: no unexplained differences',
                      '; '.join(unexpected[:3]))


def read_sd_pairs(meas_root):
    """Read the real GLOB/PROF files for one date folder off the SD-card tree."""
    root = os.path.join(meas_root, DATE_FOLDER, 'PART0000')
    if not os.path.isdir(root):
        print(f'SKIP: no reference data at {root}')
        raise SystemExit(0)
    pairs = []
    for proj in sorted(os.listdir(root)):
        pdir = os.path.join(root, proj)
        if not os.path.isdir(pdir):
            continue
        for fn in sorted(os.listdir(pdir)):
            with open(os.path.join(pdir, fn), 'rb') as fh:
                pairs.append((f'{DATE_FOLDER}/PART0000/{proj}/{fn}', fh.read()))
    return pairs


META = {
    'recorder_name': 'Catherine Ives-Yim',
    'location_label': 'Reference site',
    'postcode': 'LS1 1AA', 'lat': 53.7997, 'lng': -1.5492,
    'notes': 'a note that must survive the hop',
}


def main(meas_root):
    pairs = read_sd_pairs(meas_root)
    tmp = tempfile.mkdtemp(prefix='noisetest-')
    sys.path.insert(0, REPO)
    try:
        from noise_parser import parse_files
        from nor140_format import SPECTRAL_TABLES
        sessions = parse_files(pairs)

        # ── 1. import_sessions() end to end ───────────────────────────────────
        print('\n1. import_sessions() from real SD-card binaries')
        a = Side(os.path.join(tmp, 'a.db'))
        n = a.db.import_sessions(sessions, metadata=META)
        check(n == len(sessions) == 1, 'one session imported', f'n={n}')

        data = a.db.get_all_sessions_json()['sessions']
        check(len(data) == 1, 'one session read back')
        sess = data[0]
        check(sess['d'] == SESSION_DATE, 'session date', sess['d'])
        n_runs = len(sess['projects'])
        check(n_runs == len(sessions[0]['projects']), 'run count round-trips',
              f'{n_runs} runs')
        check(sess['name'] == 'Catherine Ives-Yim', 'recorder_name stored')
        check(sess['post'] == 'LS1 1AA', 'postcode stored')

        # Run 9 against the Nortfr-confirmed reference values from NOR140_handoff.md.
        # This exercises parse -> INSERT -> SELECT for the GLOB scalar columns;
        # an INSERT placeholder miscount or column-order slip shows up here.
        run9 = a.db.get_full_run_row(SESSION_DATE, 9)
        check(run9 is not None, 'run 9 present')
        close(run9['avg_laeq'], 70.8, 0.05, 'run 9 LAeq  = 70.8 (Nortfr)')
        close(run9['lceq'],     78.7, 0.05, 'run 9 LCeq  = 78.7 (Nortfr)')
        close(run9['lapeak'],   97.4, 0.05, 'run 9 LApeak= 97.4 (Nortfr)')
        close(run9['lcpeak'],   98.6, 0.05, 'run 9 LCpeak= 98.6 (Nortfr)')
        close(run9['la_l10'],   74.8, 0.05, 'run 9 LA10  = 74.8 (Nortfr)')
        close(run9['la_l50'],   67.1, 0.05, 'run 9 LA50  = 67.1 (Nortfr)')
        close(run9['la_l90'],   57.1, 0.05, 'run 9 LA90  = 57.1 (Nortfr)')
        check(run9['n_samples'] == 900, 'run 9 is 900 s', str(run9['n_samples']))

        # The meter's own end time (GLOB 0x22). Stored, not derived: for run 9 it
        # happens to equal start + 900 s, but 250712/24 ends at 21:27:49 — which
        # matches neither its 304 periods nor its 299 s duration, so any
        # arithmetic on start time would report the wrong figure there.
        check(run9['end_time'] == '23:42:36', 'run 9 end time = 23:42:36',
              str(run9['end_time']))
        missing_end = [p['start'] for p in sess['projects'] if not p.get('end')]
        check(not missing_end, 'every run exposes an end time', str(missing_end))
        r9 = next(p for p in sess['projects'] if p['start'] == run9['start_time'])
        check(r9['end'] == '23:42:36', "run 9 'end' reaches the display layer",
              str(r9.get('end')))

        # The end-time field is sometimes left unset. The decoder cross-checks it
        # against the stored duration rather than blacklisting 00:00:00, since a
        # run can legitimately end at midnight.
        import glob as _glob
        from nor140_format import read_end_time
        _corpus = sorted(_glob.glob(os.path.join(meas_root, '**', 'GLOB*.DAT'),
                                   recursive=True))
        _rejected = [g for g in _corpus
                     if read_end_time(open(g, 'rb').read()) is None]
        check(len(_rejected) == 2, 'end-time guard rejects exactly the 2 unset fields',
              f'{len(_rejected)} of {len(_corpus)}')
        check(all(d in r for d in ('230830', '250711') for r in [' '.join(_rejected)]),
              'the rejected pair are the known 230830/250711 runs')

        # The 1069-byte GLOB variant pairs with a 4-channel, 8-byte PROF record.
        # Read as 10 bytes the whole series is misaligned, so these assertions
        # are the guard against that regressing: the decoded profile has to
        # reproduce the GLOB scalars exactly.
        import noise_parser as parser
        import collections
        from nor140_format import prof_record_size, read_duration_s
        _b = os.path.join(meas_root, '250823', 'PART0000', 'PROJ0002')
        if os.path.isdir(_b):
            _g = open(os.path.join(_b, 'GLOB0002.DAT'), 'rb').read()
            _p = open(os.path.join(_b, 'PROF0002.DAT'), 'rb').read()
            check(prof_record_size(len(_p) - 3, 894, 1069) == 8,
                  '1069-byte variant resolves to 8-byte records')

            _d, _r = parser._parse_session_files(_g, _p)
            check(_r['n'] == 895, '8-byte run has 895 records', str(_r['n']))
            close(parser._energy_avg(_r['prof_laeq_json']), 80.28, 0.005,
                  '8-byte profile LAeq reproduces the GLOB scalar')
            close(max(_r['prof_lapeak_json']), 102.83, 0.005,
                  '8-byte profile LApeak reproduces the GLOB scalar')
            check(_r['prof_lae_json'] is None,
                  '4-channel layout stores no LAE series rather than faking one')

        # prof_record_size() weighs two signals: the duration and the GLOB
        # variant. Both are derived before either is returned, so a file whose
        # signals disagree is refused rather than guessed at. The conflict case
        # is the one that regressed — the duration branch used to return first.
        check(prof_record_size(7160, 716, 1069) is None,
              'conflicting duration and GLOB variant fails closed')
        check(prof_record_size(7160, 3000, 1069) == 8,
              'paused 1069 variant resolves by variant when duration cannot')
        check(prof_record_size(7160, 3000, None) is None,
              'unknown variant with no duration match is unresolved')
        check(prof_record_size(9000, 900, 2653) == 10,
              'normal 2653 variant resolves to 10-byte records')
        check(prof_record_size(7160, None, 1069) == 8,
              'variant alone resolves when there is no duration')

        # Every historical file must still classify, and to the same size as
        # before: an unresolved file is skipped on import, so a regression here
        # silently drops runs.
        _sizes = collections.Counter()
        for _gg in _corpus:
            _pp = _gg.replace('GLOB', 'PROF')
            if not os.path.exists(_pp):
                continue
            _gd = open(_gg, 'rb').read()
            _sizes[prof_record_size(os.path.getsize(_pp) - 3,
                                    read_duration_s(_gd), len(_gd))] += 1
        check(_sizes[None] == 0, 'no corpus file is left unclassified',
              f'{_sizes[None]} unresolved')
        check(_sizes[10] == 473 and _sizes[8] == 54,
              'corpus classifies 473 ten-byte / 54 eight-byte', str(dict(_sizes)))

        # One example is not enough to pin a channel layout, so assert it over
        # every 8-byte file in the archive: ch2 must reproduce LAFmax and ch3
        # LApeak exactly, and ch1's energy average must be the stored LAeq.
        _short = []
        for _g8 in _corpus:
            _gd = open(_g8, 'rb').read()
            if len(_gd) != 1069:
                continue
            _pf = _g8.replace('GLOB', 'PROF')
            if not os.path.exists(_pf):
                continue
            _d, _rr = parser._parse_session_files(_gd, open(_pf, 'rb').read())
            if _rr:
                _short.append((_g8, _rr))
        check(len(_short) == 54, 'all 54 short-format runs parse', str(len(_short)))
        _bad_max = [g for g, r in _short
                    if r['lafmax'] is None or r['prof_lafmax_json'] is None
                    or abs(max(r['prof_lafmax_json']) - r['lafmax']) > 0.005]
        check(not _bad_max, 'ch2 max == GLOB LAFmax in all 54', str(_bad_max[:2]))
        _bad_peak = [g for g, r in _short
                     if r['lapeak'] is None
                     or abs(max(r['prof_lapeak_json']) - r['lapeak']) > 0.005]
        check(not _bad_peak, 'ch3 max == GLOB LApeak in all 54', str(_bad_peak[:2]))
        _bad_eq = [g for g, r in _short
                   if abs(parser._energy_avg(r['prof_laeq_json']) - r['avg']) > 0.05]
        check(len(_bad_eq) <= 2,
              'ch1 energy average == GLOB LAeq in at least 52 of 54',
              f'{len(_bad_eq)} outliers')
        _no_lae = [g for g, r in _short if r['prof_lae_json'] is not None]
        check(not _no_lae, 'no 8-byte run fabricates an LAE series')

        # ch0 vs ch1 cannot be told apart by energy average — they agree to two
        # decimals. The GLOB LAF percentiles are the independent discriminator:
        # they describe the LAFspl distribution, so ch0 must track them more
        # closely than ch1 does. Without this, a ch0/ch1 swap would pass silently.
        def _pct(series, frac):
            sv = sorted(series, reverse=True)
            return sv[max(0, int(len(sv) * frac) - 1)]

        _wrong = []
        for _g8, _rr in _short:
            _c0, _c1 = _rr['prof_lafspl_json'], _rr['prof_laeq_json']
            _ref = [_rr['la_l10'], _rr['la_l50'], _rr['la_l90']]
            if any(v is None for v in _ref):
                continue
            _e0 = sum(abs(_pct(_c0, f) - v) for f, v in zip((.1, .5, .9), _ref))
            _e1 = sum(abs(_pct(_c1, f) - v) for f, v in zip((.1, .5, .9), _ref))
            if _e0 >= _e1:
                _wrong.append((_g8.split('MEAS118/')[-1], round(_e0, 2), round(_e1, 2)))
        check(not _wrong,
              'ch0 tracks the GLOB LAF percentiles better than ch1 in all 54',
              str(_wrong[:2]))

        missing = [c for c, _ in SPECTRAL_TABLES if not run9.get(c)]
        check(not missing, 'all 18 spectral tables stored', f'missing: {missing}')
        for col, _ in SPECTRAL_TABLES:
            check(len(json.loads(run9[col])) == 36, f'{col} has 36 bands')
        for col in ('prof_lafspl_json', 'prof_laeq_json', 'prof_lafmax_json',
                    'prof_lae_json', 'prof_lapeak_json'):
            check(len(json.loads(run9[col])) == 900, f'{col} has 900 samples')

        # ── 2. LA10/LA90 pooled percentile ────────────────────────────────────
        print('\n2. pooled LA10/LA90 percentile across runs')
        pooled = a.db.get_session_prof_lafspl(SESSION_DATE, [9])
        check(len(pooled) == 900, 'run 9 pools 900 samples', str(len(pooled)))
        check(pooled == json.loads(run9['prof_lafspl_json']),
              'pooling reads PROF field 0 (LAFspl), not field 1 (LAeq,1s)')
        check(pooled != json.loads(run9['prof_laeq_json']),
              'LAFspl and LAeq,1s are genuinely different series')

        multi = a.db.get_session_prof_lafspl(SESSION_DATE, [9, 10])
        check(len(multi) == 1800, 'two runs pool 1800 samples', str(len(multi)))
        every = a.db.get_session_prof_lafspl(SESSION_DATE)
        check(len(every) == sum(p['n'] for p in sess['projects']),
              'no run_numbers pools every run', f'{len(every)} samples')

        # _percentile interpolates over an ascending list: LA90 is the level
        # exceeded 90% of the time, i.e. the 10th percentile.
        s = sorted(pooled)
        la90, la50, la10 = (a.reports._percentile(s, p) for p in (10, 50, 90))
        check(la90 < la50 < la10, 'LA90 < LA50 < LA10', f'{la90} / {la50} / {la10}')

        def naive(sorted_vals, p):
            idx = (p / 100) * (len(sorted_vals) - 1)
            lo = int(idx)
            hi = min(lo + 1, len(sorted_vals) - 1)
            return round(sorted_vals[lo] + (idx - lo) *
                         (sorted_vals[hi] - sorted_vals[lo]), 1)
        for p in (10, 50, 90):
            check(a.reports._percentile(s, p) == naive(s, p),
                  f'percentile p={p} matches independent computation')
        check(a.reports._percentile([], 50) is None, 'empty input returns None')
        check(a.reports._percentile([42.0], 90) == 42.0, 'single value returns itself')

        # pct85 must count the true 1-second series, not the expanded chart profile
        # (whose windows each hold a maximum, overstating time above the threshold).
        for rn in (5, 7):
            true_series = json.loads(a.db.get_full_run_row(SESSION_DATE, rn)['prof_laeq_json'])
            expected = round(100 * sum(1 for v in true_series if v >= 85) / len(true_series), 1)
            proj = sess['projects'][rn - 1]
            st = a.reports._run_stats(proj, true_laeq=a.db.get_run_prof_laeq(SESSION_DATE, rn))
            check(st['pct85'] == expected,
                  f'run {rn}: pct85 counts the true series', f"{st['pct85']} vs {expected}")
            inflated = a.reports._expand_run(proj)
            infl = round(100 * sum(1 for v in inflated if v >= 85) / len(inflated), 1)
            check(infl >= st['pct85'],
                  f'run {rn}: the old chart-profile count was higher ({infl} vs {st["pct85"]})')

        p9 = a.reports._percentile(sorted(a.db.get_session_prof_lafspl(SESSION_DATE, [9])), 90)
        p10 = a.reports._percentile(sorted(a.db.get_session_prof_lafspl(SESSION_DATE, [10])), 90)
        pooled_910 = a.reports._percentile(sorted(multi), 90)
        check(abs(pooled_910 - (p9 + p10) / 2) > 1e-9,
              'pooled LA10 differs from the mean of per-run LA10s',
              f'pooled={pooled_910}, mean-of-runs={round((p9 + p10) / 2, 1)}')

        # ── 3. delete_session() cascade ───────────────────────────────────────
        print('\n3. delete_session() cascade')
        aid = a.assess.create_assessment('Cascade test', standard='bs4142')
        loc = a.assess.add_assessment_location(aid, 'Loc A')
        a.assess.assign_runs(aid, loc, [(SESSION_DATE, 9), (SESSION_DATE, 10)])
        a.db.save_weather(SESSION_DATE, {
            'wind_speed': 7.2, 'wind_dir': 210.5, 'temp_min': 11.0,
            'temp_max': 19.4, 'precip': 0.2, 'hourly_json': '{}'})

        def n_of(side, table, where='', *p):
            return side.sql(f'SELECT COUNT(*) c FROM {table} {where}', *p)[0]['c']

        check(n_of(a, 'assessment_runs') == 2, 'two runs assigned')
        check(n_of(a, 'runs') == n_runs, 'runs present before delete', str(n_runs))

        a.db.delete_session(SESSION_DATE)

        check(n_of(a, 'sessions', 'WHERE date=?', SESSION_DATE) == 0, 'session row deleted')
        check(n_of(a, 'runs') == 0, 'runs cascaded via the sessions FK')
        check(n_of(a, 'assessment_runs') == 0,
              'assessment_runs deleted by hand (session_date is not a real FK)')
        check(n_of(a, 'assessments') == 1,
              'the assessment itself survives — only its run assignments go')
        check(n_of(a, 'assessment_locations') == 1, 'assessment locations survive')
        check(n_of(a, 'weather', 'WHERE date=?', SESSION_DATE) == 1,
              'weather row is intentionally NOT deleted (per-date data)')

        # ── 4. sync-shaped round trip ─────────────────────────────────────────
        print('\n4. peer-sync round trip preserves metadata and weather')
        a.db.import_sessions(sessions, metadata=META)
        a.db.save_weather(SESSION_DATE, {
            'wind_speed': 7.2, 'wind_dir': 210.5, 'temp_min': 11.0,
            'temp_max': 19.4, 'precip': 0.2, 'hourly_json': '{}'})
        payload = a.db.get_sessions_since('1970-01-01T00:00:00')
        check(len(payload) == 1, 'one session in the sync payload')
        check(payload[0]['wx'] is not None, 'weather included in payload')
        check(payload[0]['notes'] == META['notes'], 'notes included in payload')
        check(payload[0]['projects'][8].get('spec_lfeq') is not None,
              'payload carries spectral arrays (full=True)')

        b = Side(os.path.join(tmp, 'b.db'))
        b.db.import_sessions(payload)          # no metadata kwarg — as the peer does

        psess = b.db.get_all_sessions_json()['sessions'][0]
        check(psess['name'] == 'Catherine Ives-Yim', 'recorder_name survived the hop')
        check(psess['loc'] == 'Reference site', 'location_label survived')
        check(psess['post'] == 'LS1 1AA', 'postcode survived')
        check(psess['notes'] == META['notes'], 'notes survived')
        close(psess['lat'], 53.7997, 1e-9, 'lat survived')
        close(psess['lng'], -1.5492, 1e-9, 'lng survived')
        check(psess['wx'] is not None, 'weather survived the hop')
        close(psess['wx']['ws'], 7.2, 1e-9, 'wind speed survived')
        close(psess['wx']['tx'], 19.4, 1e-9, 'temp max survived')

        prun9 = b.db.get_full_run_row(SESSION_DATE, 9)
        close(prun9['avg_laeq'], 70.8, 0.05, 'peer run 9 LAeq intact')
        close(prun9['la_l90'],   57.1, 0.05, 'peer run 9 LA90 intact')
        check(len(json.loads(prun9['spec_lfeq'])) == 36, 'peer run 9 spectral table intact')
        check(len(json.loads(prun9['prof_lafspl_json'])) == 900,
              'peer run 9 PROF series intact — xlsx export works on the peer')
        check(b.db.get_session_prof_lafspl(SESSION_DATE, [9]) ==
              a.db.get_session_prof_lafspl(SESSION_DATE, [9]),
              'pooled LAFspl identical on both sides')

        # The meter's end time must survive the hop. This regressed once because
        # _run_to_dict exported it as 'end' while import_sessions read
        # 'end_time', so the peer silently stored NULL. Run 1 is the case that
        # exposes it: 83 records but an 82 s duration, so any arithmetic
        # reconstruction lands a second late, at 12:09:53 rather than 12:09:52.
        arun1 = a.db.get_full_run_row(SESSION_DATE, 1)
        prun1 = b.db.get_full_run_row(SESSION_DATE, 1)
        check(arun1['end_time'] == '12:09:52', 'source run 1 end time',
              str(arun1['end_time']))
        check(prun1['end_time'] == arun1['end_time'],
              'meter end time survives peer sync', str(prun1['end_time']))
        check(prun1['n_samples'] == 83 and prun1['duration_s'] == 82,
              'run 1 is the count-minus-one case that arithmetic gets wrong')
        pdisp = next(pp for pp in psess['projects'] if pp['run_number'] == 1)
        check(pdisp['end'] == '12:09:52', "peer displays the meter's end time",
              str(pdisp.get('end')))
        check(pdisp.get('end_time') == '12:09:52',
              'peer re-exports end_time for onward sync')

        # ── 4b. assessment links survive a run-number shift ───────────────────
        print('\n4b. assessment links are keyed on the stable run identity')
        aid = a.assess.create_assessment('Shift test', standard='bs4142')
        lid = a.assess.add_assessment_location(aid, 'Boundary')
        a.assess.assign_runs(aid, lid, [(SESSION_DATE, 5)])
        before = a.assess.get_assessment_detail(aid)['locations'][0]['runs'][0]
        check(before['run_number'] == 5, 'link created against run 5')
        pinned_start = before['start_time']
        ar_src = a.sql('SELECT source_file FROM assessment_runs')[0]['source_file']
        check(ar_src is not None, 'link stored the stable source_file', str(ar_src))

        # Simulate what a re-import does when a previously unparseable run is
        # recovered: an extra run appears earlier in the session and every later
        # run shifts up one. Keyed on run_number alone, the link would silently
        # follow the number onto a different measurement.
        sid = a.sql('SELECT id FROM sessions WHERE date=?', SESSION_DATE)[0]['id']
        # Two steps: a single +1 pass collides with UNIQUE(session_id, run_number)
        # as SQLite rewrites row by row.
        a.exec('UPDATE runs SET run_number = run_number + 1000 '
               'WHERE session_id=? AND run_number >= 3', sid)
        a.exec('UPDATE runs SET run_number = run_number - 999 '
               'WHERE session_id=? AND run_number >= 1000', sid)
        a.exec("INSERT INTO runs (session_id, run_number, source_file, start_time, "
               "n_samples, step, avg_laeq) VALUES (?,3,'PROJ9999','20:00:00',60,1,50.0)", sid)

        after = a.assess.get_assessment_detail(aid)['locations'][0]['runs'][0]
        check(after['start_time'] == pinned_start,
              'link still resolves to the same physical run after the shift',
              f"{pinned_start} -> {after['start_time']}")
        check(after['run_number'] == 6,
              'and follows it to its new position', str(after['run_number']))
        moved = a.sql('SELECT start_time FROM runs WHERE session_id=? AND run_number=5',
                      sid)[0]['start_time']
        check(moved != pinned_start,
              'run 5 is now a different measurement — what the old key would have returned',
              f'{moved} vs {pinned_start}')

        # A peer on the older schema sends no source_file; that must not erase ours.
        a.sync.apply_sync_event('assessment_run', 'upsert', {
            'id': a.sql('SELECT id FROM assessment_runs')[0]['id'],
            'assessment_id': aid, 'location_id': lid, 'session_date': SESSION_DATE,
            'run_number': 6, 'conditions': 'dry', 'notes': ''})
        kept = a.sql('SELECT source_file FROM assessment_runs')[0]['source_file']
        check(kept == ar_src, "an old peer's payload does not erase source_file",
              str(kept))

        # The two write failures the read-side fix alone did not prevent: with
        # the upsert still keyed on run_number, re-assigning the same physical
        # run at its new number inserted a duplicate, and assigning whatever had
        # taken the old number silently rebound the link.
        a.assess.assign_runs(aid, lid, [(SESSION_DATE, 6)])
        rows = a.sql('SELECT run_number, source_file FROM assessment_runs')
        check(len(rows) == 1, 're-assigning the same run at its new number is not a duplicate',
              str([(r['run_number'], r['source_file']) for r in rows]))
        check(rows[0]['source_file'] == ar_src, 'and still points at the same measurement')
        check(rows[0]['run_number'] == 6, 'with run_number refreshed to the new position',
              str(rows[0]['run_number']))

        a.assess.assign_runs(aid, lid, [(SESSION_DATE, 5)])
        srcs = sorted(r['source_file'] for r in
                      a.sql('SELECT source_file FROM assessment_runs'))
        check(ar_src in srcs, 'assigning the old number does not rebind the original link',
              str(srcs))
        check(len(srcs) == 2, 'it creates a separate link for the other measurement',
              str(srcs))

        # The picker must mark the run that is actually linked, not whichever
        # run inherited the number — that is what led users into the above.
        pool = a.assess.get_all_runs_for_assessment(aid)
        marked = {r['source_file'] for r in pool
                  if r['ar_id'] and r['session_date'] == SESSION_DATE}
        check(ar_src in marked, 'picker marks the linked measurement', str(sorted(marked)))

        # The session card and the assessment detail must not disagree.
        card = [x for x in a.db.get_all_sessions_json()['sessions']
                if x['d'] == SESSION_DATE][0]
        labelled = {p['run_number'] for p in card['projects'] if p.get('assess')}
        detail_nums = {r['run_number'] for r in
                       a.assess.get_assessment_detail(aid)['locations'][0]['runs']}
        check(labelled == detail_nums,
              'session card and assessment detail agree on which runs are linked',
              f'card {sorted(labelled)} vs detail {sorted(detail_nums)}')

        a.assess.delete_assessment(aid)
        a.exec('DELETE FROM runs WHERE source_file=?', 'PROJ9999')
        a.exec('UPDATE runs SET run_number = run_number + 1000 '
               'WHERE session_id=? AND run_number >= 4', sid)
        a.exec('UPDATE runs SET run_number = run_number - 1001 '
               'WHERE session_id=? AND run_number >= 1000', sid)

        # ── 4ante. the import path and the shared decoders agree ─────────────
        print('\n4ante. no divergent copies of the binary decoders')
        # A duplicated decoder fails silently: the reference comparison exercises
        # the export path, while a different function writes the database. That
        # is how the -20 marker survived, and how impulse spectra above 130 dB
        # were dropped on import while the backfill kept them. This asserts the
        # two paths still agree over every table in the archive.
        import noise_parser as _np
        from nor140_format import (read_glob_spectrum as _rgs,
                                   read_glob_scalars as _rsc,
                                   read_start_datetime as _rsd)
        _div = _scal = _dt = 0
        for _gg in _corpus:
            _bd = open(_gg, 'rb').read()
            for _col, _off in a.db.SPECTRAL_TABLES:
                if _np._read_glob_spectrum(_bd, _off) != _rgs(_bd, _off):
                    _div += 1
            if _np._read_glob_scalars(_bd) != _rsc(_bd):
                _scal += 1
            if _rsd(_bd)[0] is None and len(_bd) > 0x20:
                _dt += 1
        check(_div == 0, 'spectrum decode identical on both paths', f'{_div} divergent')
        check(_scal == 0, 'scalar decode identical on both paths', f'{_scal} divergent')
        check(_dt == 0, 'every archive file yields a start datetime', f'{_dt} failed')

        # Impulse tables legitimately exceed CAP_LAEQ and must survive.
        _imp = 0
        for _gg in _corpus:
            _bd = open(_gg, 'rb').read()
            for _col, _off in a.db.SPECTRAL_TABLES:
                _t = _rgs(_bd, _off)
                if _t and max(_t) > 130.0:
                    _imp += 1
        check(_imp >= 3, 'spectra above the LAeq cap are kept, not discarded',
              f'{_imp} tables above 130 dB')

        # ── 4bis. the -20 dB no-data marker never reaches a report ────────────
        print('\n4bis. no-data marker is dropped at the decode boundary')
        r1 = a.db.get_full_run_row(SESSION_DATE, 1)
        check(r1['n_samples'] == 83, 'run 1 is short enough to lack a 0.1% percentile')
        check(r1['la_l01'] is None, 'unrecorded percentile stored as NULL, not -20',
              repr(r1['la_l01']))
        check(r1['la_l90'] is not None, 'recorded percentiles are unaffected',
              repr(r1['la_l90']))

        # A stored -20 must not survive into the BS 4142 figures: fed through as
        # an LA90 it produces a rating-level difference of about +81 dB.
        a.exec('UPDATE runs SET la_l90=-20.0, la_l50=-20.0 WHERE run_number=2')
        cleared = a.db.clear_sentinel_scalars()
        check(cleared.get('la_l90') == 1, 'maintenance helper clears a stored marker',
              str(cleared))
        check(a.db.get_full_run_row(SESSION_DATE, 2)['la_l90'] is None,
              'and leaves NULL behind')

        aid2 = a.assess.create_assessment('Sentinel', standard='bs4142')
        lid2 = a.assess.add_assessment_location(aid2, 'B')
        a.assess.assign_runs(aid2, lid2, [(SESSION_DATE, 2)])
        rd = a.assess.prepare_assessment_report_data(aid2)['locations'][0]['runs'][0]
        check(rd.get('la90') is None or rd['la90'] > -19.99,
              'report data carries no no-data marker', repr(rd.get('la90')))
        a.assess.delete_assessment(aid2)

        # ── 4d. the assign route, end to end ──────────────────────────────────
        print('\n4d. assignment over HTTP, not just the data layer')
        # The suite only ever called assign_runs() directly, so a 500 in the
        # route sailed past every check: the endpoint built three-element tuples
        # and handed them to a helper that unpacked two. assign_runs had already
        # committed, so the row landed, the request failed, and the peer was
        # never told.
        #
        # Its own Side, because constructing one rebinds sys.modules — the route
        # resolves assign_runs through whichever binding is current, so sharing a
        # Side with an earlier section writes to the wrong database.
        import importlib as _il
        rt = Side(os.path.join(tmp, 'route.db'))
        rt.db.import_sessions(sessions, metadata=META)
        _app = _il.import_module('noise_app').app
        _app.config['TESTING'] = True
        _client = _app.test_client()
        with _client.session_transaction() as _sess:
            _sess['user'] = 'test'
            _sess['logged_in'] = True
            _sess['_csrf_token'] = SUITE_CSRF   # every POST below is CSRF-checked

        aid4 = rt.assess.create_assessment('Route test')
        lid4 = rt.assess.add_assessment_location(aid4, 'L')
        src5 = rt.sql('SELECT r.source_file FROM runs r JOIN sessions s ON s.id=r.session_id '
                      'WHERE s.date=? AND r.run_number=5', SESSION_DATE)[0]['source_file']
        _resp = _client.post(f'/api/assessments/{aid4}/assign', headers=CSRF_HDR, json={
            'location_id': lid4,
            'runs': [{'date': SESSION_DATE, 'run_number': 5, 'source_file': src5}]})
        check(_resp.status_code == 200, 'assign route returns 200, not 500',
              f'{_resp.status_code}: {_resp.get_data(as_text=True)[:120]}')
        check(_resp.get_json().get('assigned') == 1, 'route reports one assignment',
              str(_resp.get_json()))

        # A stale page: the number moved between render and POST, but the stable
        # key the page carried still names the right measurement.
        sid4 = rt.sql('SELECT id FROM sessions WHERE date=?', SESSION_DATE)[0]['id']
        rt.exec('UPDATE runs SET run_number=run_number+1000 WHERE session_id=? AND run_number>=3', sid4)
        rt.exec('UPDATE runs SET run_number=run_number-999 WHERE session_id=? AND run_number>=1000', sid4)
        _resp = _client.post(f'/api/assessments/{aid4}/assign', headers=CSRF_HDR, json={
            'location_id': lid4,
            'runs': [{'date': SESSION_DATE, 'run_number': 5, 'source_file': src5}]})
        check(_resp.status_code == 200, 'a stale run number still assigns cleanly',
              str(_resp.status_code))
        _srcs = [r['source_file'] for r in
                 rt.sql('SELECT source_file FROM assessment_runs WHERE assessment_id=?', aid4)]
        check(_srcs == [src5], 'the stable key won over the stale position', str(_srcs))

        # A source_file that is not in this session is refused, not guessed at.
        _resp = _client.post(f'/api/assessments/{aid4}/assign', headers=CSRF_HDR, json={
            'location_id': lid4,
            'runs': [{'date': SESSION_DATE, 'run_number': 5, 'source_file': 'PROJ7777'}]})
        check(_resp.status_code == 400, 'a foreign source_file is rejected',
              str(_resp.status_code))

        # ── 4c. migration audit on hostile inputs ─────────────────────────────
        print('\n4c. migration audit flags links it cannot safely migrate')
        aid3 = a.assess.create_assessment('Audit test')
        lid3 = a.assess.add_assessment_location(aid3, 'L')
        a.assess.assign_runs(aid3, lid3, [(SESSION_DATE, 2)])
        clean = a.db.audit_assessment_run_keys()
        check(not any(clean.values()), 'a healthy table audits clean',
              str({k: len(v) for k, v in clean.items()}))

        # A link whose stable key names a run that is not there — what an
        # already-shifted database looks like after a naive backfill.
        a.exec("UPDATE assessment_runs SET source_file='PROJ8888' "
               'WHERE assessment_id=?', aid3)
        rep = a.db.audit_assessment_run_keys()
        check(len(rep['unmatched']) == 1, 'audit reports an unresolvable stable key',
              str(len(rep['unmatched'])))

        # A legacy row that never received one.
        a.exec('UPDATE assessment_runs SET source_file=NULL WHERE assessment_id=?', aid3)
        rep = a.db.audit_assessment_run_keys()
        check(len(rep['null_source_file']) == 1, 'audit reports a missing stable key',
              str(len(rep['null_source_file'])))
        # ...and such a row must still be assignable without creating a second.
        a.assess.assign_runs(aid3, lid3, [(SESSION_DATE, 2)])
        n_legacy = len(a.sql('SELECT id FROM assessment_runs WHERE assessment_id=?', aid3))
        check(n_legacy == 1, 'a legacy NULL-key link is updated, not duplicated',
              str(n_legacy))
        a.assess.delete_assessment(aid3)

        # The positional UNIQUE must be gone, replaced by the stable one.
        ddl = a.sql("SELECT sql FROM sqlite_master WHERE name='assessment_runs'")[0]['sql']
        check('UNIQUE(assessment_id, session_date, run_number)' not in ddl,
              'the positional UNIQUE is dropped')
        idx = a.sql("SELECT name FROM sqlite_master WHERE type='index' "
                    "AND name='idx_assessment_runs_stable'")
        check(len(idx) == 1, 'the stable partial unique index exists')

        # The client half of the stable key is JavaScript, so nothing above can
        # exercise it — which is precisely how it was silently lost once: an edit
        # that never landed, while the brief claimed the protection existed.
        _tpl = open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 'templates', 'assessments.html'), encoding='utf-8').read()
        _keyline = [l for l in _tpl.split('\n') if 'const key = ' in l]
        # WP8: the key also carries the instrument serial — two meters can hold
        # a same-numbered run (even the same PROJ folder) on one date.
        check(len(_keyline) == 1 and 'r.source_file' in _keyline[0]
              and 'r.instrument_serial' in _keyline[0],
              'the pool key carries source_file and the serial, not just the position',
              _keyline[0].strip() if _keyline else 'not found')
        check('const [date, rn, srcFile, serial] = key.split' in _tpl,
              'and the click handler destructures all four')
        check('source_file: srcFile || null' in _tpl,
              'and posts source_file to the server')
        check("serial: serial || ''" in _tpl,
              'and posts the serial to the server')

        # ── 4e. migration from hostile starting schemas ───────────────────────
        print('\n4e. migration on databases unlike the ones it was written against')
        import sqlite3 as _sq, importlib as _il2
        _BASE = ("CREATE TABLE assessments(id INTEGER PRIMARY KEY, name TEXT);"
                 "CREATE TABLE assessment_locations(id INTEGER PRIMARY KEY, assessment_id INTEGER);"
                 "CREATE TABLE sessions(id INTEGER PRIMARY KEY, date TEXT UNIQUE);"
                 "CREATE TABLE runs(id INTEGER PRIMARY KEY, session_id INTEGER,"
                 "  run_number INTEGER, source_file TEXT);")

        def _build(name, extra):
            path = os.path.join(tmp, name)
            cx = _sq.connect(path); cx.executescript(_BASE + extra); cx.commit(); cx.close()
            return path

        def _migrate_only(path):
            """Run the migration against `path` without disturbing the Side bindings."""
            prev = os.environ.get('NOISE_DB_PATH')
            os.environ['NOISE_DB_PATH'] = path
            for _m in ('noise_db', 'reports_db', 'assessments_db', 'sync_db', 'reports'):
                sys.modules.pop(_m, None)
            try:
                _nd = _il2.import_module('noise_db')
                _nd.init_db()
                return None
            except Exception as exc:
                return exc
            finally:
                if prev:
                    os.environ['NOISE_DB_PATH'] = prev
                for _m in ('noise_db', 'reports_db', 'assessments_db', 'sync_db', 'reports'):
                    sys.modules.pop(_m, None)

        # (a) a database predating the source_file column. The index used to be
        # created in the schema script, before the ALTER that adds the column.
        _p = _build('old_schema.db', """
            CREATE TABLE assessment_runs(id INTEGER PRIMARY KEY,
              assessment_id INTEGER NOT NULL REFERENCES assessments(id) ON DELETE CASCADE,
              location_id INTEGER, session_date TEXT NOT NULL, run_number INTEGER NOT NULL,
              conditions TEXT, notes TEXT,
              UNIQUE(assessment_id, session_date, run_number));
            INSERT INTO sessions(id,date) VALUES(1,'2026-08-12');
            INSERT INTO runs(id,session_id,run_number,source_file) VALUES(1,1,1,'PROJ0001');
            INSERT INTO assessments(id,name) VALUES(1,'A');
            INSERT INTO assessment_runs(id,assessment_id,session_date,run_number)
              VALUES(1,1,'2026-08-12',1);""")
        _err = _migrate_only(_p)
        check(_err is None, 'pre-source_file schema migrates', repr(_err))
        _cx = _sq.connect(_p)
        check(_cx.execute('SELECT source_file FROM assessment_runs').fetchone()[0] == 'PROJ0001',
              'and the existing link is backfilled')
        check('UNIQUE(assessment_id, session_date, run_number)' not in
              _cx.execute("SELECT sql FROM sqlite_master WHERE name='assessment_runs'").fetchone()[0],
              'and the positional UNIQUE is dropped')
        _cx.close()

        # (b) duplicated stable keys must refuse, not fail with a bare
        # IntegrityError naming nothing.
        _p = _build('dupe_keys.db', """
            CREATE TABLE assessment_runs(id INTEGER PRIMARY KEY,
              assessment_id INTEGER NOT NULL REFERENCES assessments(id) ON DELETE CASCADE,
              location_id INTEGER, session_date TEXT NOT NULL, run_number INTEGER,
              source_file TEXT, conditions TEXT, notes TEXT);
            INSERT INTO assessments(id,name) VALUES(1,'A');
            INSERT INTO assessment_runs(id,assessment_id,session_date,run_number,source_file)
              VALUES(1,1,'2026-08-12',5,'PROJ0005'),(2,1,'2026-08-12',6,'PROJ0005');""")
        _err = _migrate_only(_p)
        check(type(_err).__name__ == 'MigrationUnsafe',
              'duplicated stable keys refuse the migration', repr(_err))
        check('audit_assessment_run_keys' in str(_err),
              'and the message says how to resolve it')

        # (c) two runs sharing one source_file make the key ambiguous.
        _p = _build('dupe_runs.db', """
            CREATE TABLE assessment_runs(id INTEGER PRIMARY KEY,
              assessment_id INTEGER NOT NULL REFERENCES assessments(id) ON DELETE CASCADE,
              location_id INTEGER, session_date TEXT NOT NULL, run_number INTEGER,
              source_file TEXT, conditions TEXT, notes TEXT);
            INSERT INTO sessions(id,date) VALUES(1,'2026-08-12');
            INSERT INTO runs(id,session_id,run_number,source_file)
              VALUES(1,1,1,'PROJ0001'),(2,1,2,'PROJ0001');""")
        _err = _migrate_only(_p)
        check(type(_err).__name__ == 'MigrationUnsafe',
              'duplicated run identities refuse the migration', repr(_err))

        # ── 4f. run identity under partial and shifted imports ────────────────
        print('\n4f. run identity under partial and shifted imports')
        # Runs used to be keyed on their position in the payload, so a ZIP of
        # just PROJ0004 became "run 1" and overwrote PROJ0001; a re-import with
        # the first file unreadable shifted every run down one and then died on
        # the source_file index. The key is now (session, source_file).
        import copy as _copy
        import zipfile as _zf
        import io as _io
        from nor140_format import PROF_RECORD_OFFSET, PROF_RECORD_SIZE
        from noise_parser import parse_zip as _parse_zip

        def _runs(side):
            return [(r['run_number'], r['source_file'], r['avg_laeq'], r['start_time'])
                    for r in side.sql(
                        'SELECT r.run_number, r.source_file, r.avg_laeq, r.start_time '
                        'FROM runs r JOIN sessions s ON s.id=r.session_id '
                        'WHERE s.date=? ORDER BY r.run_number', SESSION_DATE)]

        def _sess_row(side):
            return side.sql('SELECT run_count, avg_laeq, max_laeq FROM sessions WHERE date=?',
                            SESSION_DATE)[0]

        pi = Side(os.path.join(tmp, 'identity.db'))
        full = _copy.deepcopy(sessions)
        pi.db.import_sessions(full, metadata=META)
        base_runs = _runs(pi)
        base_sess = _sess_row(pi)
        N = len(base_runs)
        check(N >= 5 and base_runs[0][1] == 'PROJ0001',
              f'baseline: {N} runs, run 1 is PROJ0001', str(base_runs[:2]))
        check([r[0] for r in base_runs] == list(range(1, N + 1)),
              'baseline numbering is 1..n')
        check(base_sess['run_count'] == N, 'baseline session run_count', str(dict(base_sess)))

        # Link an assessment to PROJ0005 so we can watch it through the churn.
        aid_f = pi.assess.create_assessment('Identity')
        lid_f = pi.assess.add_assessment_location(aid_f, 'L')
        pi.assess.assign_runs(aid_f, lid_f, [(SESSION_DATE, 5, 'PROJ0005')])
        link_start = base_runs[4][3]

        def _linked():
            runs_ = pi.assess.get_assessment_detail(aid_f)['locations'][0]['runs']
            return [(r['run_number'], r['start_time']) for r in runs_]

        check(_linked() == [(5, link_start)], 'link on run 5 = PROJ0005')

        # (a) A partial upload: just the PROJ0004 folder, as the upload page
        # allows. No date folder in the path, so it cannot be a complete date.
        partial_pairs = [(p.split('/', 2)[2], d) for p, d in pairs if '/PROJ0004/' in p]
        check(len(partial_pairs) == 2 and partial_pairs[0][0].startswith('PROJ0004/'),
              'partial upload is the two PROJ0004 files', str([p for p, _ in partial_pairs]))
        partial = parse_files(partial_pairs)
        check(len(partial) == 1 and partial[0]['complete_date'] is False,
              'a PROJ-level upload is not a complete date', str(partial[0].get('complete_date')))
        check(partial[0]['projects'][0]['source_file'] == 'PROJ0004',
              'and names its source_file', str(partial[0]['projects'][0]['source_file']))
        pi.db.import_sessions(partial, metadata=META)
        after = _runs(pi)
        check(after == base_runs,
              'partial upload of PROJ0004 leaves every stored run exactly as it was',
              f'{after[:4]} vs {base_runs[:4]}')
        check(dict(_sess_row(pi)) == dict(base_sess),
              'session aggregates still describe all stored runs, not the upload',
              f'{dict(_sess_row(pi))} vs {dict(base_sess)}')

        # (b) A shifted re-import: the first file failed to parse this time, so
        # the payload starts at PROJ0002. Positionally that rewrote run 1's
        # source_file to PROJ0002 while run 2 still held it -> IntegrityError.
        shifted = _copy.deepcopy(full)
        shifted[0]['projects'] = shifted[0]['projects'][1:]
        shifted[0]['complete_date'] = False
        try:
            pi.db.import_sessions(shifted, metadata=META)
            err = None
        except Exception as e:   # noqa: BLE001 — the defect was an IntegrityError
            err = e
        check(err is None, 'a shifted partial re-import succeeds', repr(err))
        check(_runs(pi) == base_runs, 'and changes nothing (every run already stored)')
        check(_linked() == [(5, link_start)],
              'assessment link untouched by the shifted re-import', str(_linked()))

        # (c) Complete date, but PROJ0001 is listed as skipped (present on the
        # card, unreadable today): it must survive the re-import.
        protected = _copy.deepcopy(shifted)
        protected[0]['complete_date'] = True
        protected[0]['skipped_files'] = ['PROJ0001']
        pi.db.import_sessions(protected, metadata=META)
        check(_runs(pi) == base_runs,
              'a complete-date re-import keeps a run whose file was merely unreadable')

        # (d) Complete date and PROJ0001 genuinely gone from the card.
        gone = _copy.deepcopy(shifted)
        gone[0]['complete_date'] = True
        gone[0]['skipped_files'] = []
        pi.db.import_sessions(gone, metadata=META)
        after = _runs(pi)
        check(len(after) == N - 1 and after[0][1] == 'PROJ0002',
              'complete-date re-import deletes the run that is no longer on the card',
              str(after[:2]))
        check([r[0] for r in after] == list(range(1, N)),
              'remaining runs are renumbered 1..n-1 in stored order',
              str([r[0] for r in after]))
        check(_linked() == [(4, link_start)],
              'assessment link follows PROJ0005 to its new number', str(_linked()))
        sr = _sess_row(pi)
        check(sr['run_count'] == N - 1, 'session run_count recomputed after the delete',
              str(dict(sr)))
        # WP5/D2: duration-weighted, not the equal-weight mean this check used
        # to recompute. The runs of this date differ in length, so the two
        # formulas give genuinely different answers (~1 dB apart here).
        _rows = pi.sql(
            'SELECT r.avg_laeq, r.duration_s, r.n_samples FROM runs r '
            'JOIN sessions s ON s.id=r.session_id WHERE s.date=? ORDER BY r.run_number',
            SESSION_DATE)
        _w = [(r['avg_laeq'], r['duration_s'] or r['n_samples']) for r in _rows]
        exp_avg = _rhu(10 * __import__('math').log10(
            sum(w * 10 ** (v / 10) for v, w in _w) / sum(w for _, w in _w)), 2)
        check(sr['avg_laeq'] == exp_avg, 'session avg recomputed from the remaining runs',
              f'{sr["avg_laeq"]} vs {exp_avg}')
        _flat = round(10 * __import__('math').log10(
            sum(10 ** (r[2] / 10) for r in after) / len(after)), 2)
        check(abs(_flat - exp_avg) > 0.05,
              'and the equal-weight mean it replaced is a different number',
              f'weighted={exp_avg}, equal-weight={_flat}')

        # (e) PROJ0001 comes back (full re-import): restored at run 1, the rest
        # move down, the link follows again.
        restore = _copy.deepcopy(full)
        restore[0]['complete_date'] = True
        pi.db.import_sessions(restore, metadata=META)
        check(_runs(pi) == base_runs, 'a full re-import restores the original numbering')
        check(_linked() == [(5, link_start)], 'and the link is back at run 5')
        check(dict(_sess_row(pi)) == dict(base_sess), 'session aggregates back to baseline')

        # (f) Legacy payload with no source_file at all (pre-column JSON export,
        # or a peer on the old schema): positional, no duplicates, and the
        # stored identities are not erased.
        legacy = _copy.deepcopy(full)
        for p_ in legacy[0]['projects']:
            p_.pop('source_file', None)
        legacy[0].pop('complete_date', None)
        pi.db.import_sessions(legacy, metadata=META)
        check(_runs(pi) == base_runs,
              'a legacy payload without source_file neither duplicates nor re-keys runs',
              str(_runs(pi)[:3]))

        # (g) Peer payloads carry no complete_date, so a pull can only add or
        # refresh — never delete — on the receiving side.
        peer_payload = pi.db.get_sessions_since('1970-01-01T00:00:00')
        check(all('complete_date' not in s for s in peer_payload),
              'sync payload does not claim to be a complete date')

        # ── skipped-file report ──
        def _zip(entries):
            buf = _io.BytesIO()
            with _zf.ZipFile(buf, 'w') as z:
                for name, data in entries:
                    z.writestr(name, data)
            return buf.getvalue()

        good1 = [(p, d) for p, d in pairs if '/PROJ0001/' in p]
        p2 = [(p, d) for p, d in pairs if '/PROJ0002/' in p]
        p2_glob = next(d for p, d in p2 if 'GLOB' in p.upper())
        p2_prof = next(d for p, d in p2 if 'PROF' in p.upper())
        p2_glob_path = next(p for p, _ in p2 if 'GLOB' in p.upper())
        p2_prof_path = next(p for p, _ in p2 if 'PROF' in p.upper())

        # Corrupt only the LAFspl channel (slot 0) of every record: the old check
        # looked at LAeq (slot 1) alone and would have let this through.
        body = bytearray(p2_prof[PROF_RECORD_OFFSET:])
        for off in range(0, len(body) - PROF_RECORD_SIZE + 1, PROF_RECORD_SIZE):
            body[off:off + 2] = b'\xff\xff'
        corrupt_prof = p2_prof[:PROF_RECORD_OFFSET] + bytes(body)
        rep = _parse_zip(_zip(good1 + [(p2_glob_path, p2_glob), (p2_prof_path, corrupt_prof)]))
        check(len(rep) == 1 and len(rep[0]['projects']) == 1,
              'corrupt PROJ0002 is not imported as a run', str(len(rep[0]['projects'])))
        check(len(rep.skipped) == 1 and 'PROJ0002' in rep.skipped[0]['path'],
              'parse report names the skipped pair', str(rep.skipped))
        check('dB' in rep.skipped[0]['reason'] and '140' in rep.skipped[0]['reason'],
              'and says why (level above 140 dB)', rep.skipped[0]['reason'])
        check(rep.skipped[0]['date'] == SESSION_DATE, 'skipped pair carries its date')
        check(rep[0]['complete_date'] is True and rep[0]['skipped_files'] == ['PROJ0002'],
              'session marks PROJ0002 as present-but-unreadable',
              f"{rep[0]['complete_date']} {rep[0]['skipped_files']}")
        rep2 = _parse_zip(_zip(good1 + [(p2_glob_path, p2_glob)]))
        check(len(rep2.skipped) == 1 and 'only the GLOB' in rep2.skipped[0]['reason'],
              'a lone GLOB file is reported, not ignored', str(rep2.skipped))
        rep3 = _parse_zip(_zip([(p.split('/')[-1], d) for p, d in good1]))
        check(len(rep3) == 1 and rep3[0]['complete_date'] is False and not rep3.skipped,
              'a flat DAT pair parses as a partial upload with nothing skipped')

        # Routes: /import reports the skips; /upload tells the user.
        import importlib as _il3
        import webauth as _wa
        for _m in ('noise_app', 'peer_client'):
            sys.modules.pop(_m, None)
        _app2 = _il3.import_module('noise_app')
        _app2._auto_fetch_weather = lambda *_a, **_k: None
        _app2.app.config['TESTING'] = True
        _c2 = _app2.app.test_client()
        _wa.IMPORT_KEY = 'test-key'
        corrupt_zip = _zip(good1 + [(p2_glob_path, p2_glob), (p2_prof_path, corrupt_prof)])
        r = _c2.post('/import', data={'file': (_io.BytesIO(corrupt_zip), 'x.zip')},
                     headers={'X-Import-Key': 'test-key'},
                     content_type='multipart/form-data')
        check(r.status_code == 200, '/import accepts the ZIP', str(r.status_code))
        js = r.get_json()
        check(js.get('imported') == 1 and len(js.get('skipped', [])) == 1,
              '/import response lists the skipped pair', str(js)[:200])
        # It went into the identity DB (the module bound to it). The ZIP held the
        # date folder with two PROJ folders, so it is a complete date: the other
        # eight runs are gone from the card and are deleted — but PROJ0002, which
        # was present and merely unreadable, survives alongside PROJ0001.
        check([r[1] for r in _runs(pi)] == ['PROJ0001', 'PROJ0002'],
              'complete-date import over /import deletes absent runs but keeps the skipped one',
              str([r[1] for r in _runs(pi)]))
        _wa.IMPORT_KEY = ''
        pi.db.import_sessions(restore, metadata=META)
        check(_runs(pi) == base_runs, 'full re-import restores the date again')

        with _c2.session_transaction() as _s2:
            _s2['user'] = 'test'
            _s2['logged_in'] = True
            _s2['_csrf_token'] = SUITE_CSRF   # WP2: every POST is CSRF-checked
        r = _c2.post('/upload', data={'file': (_io.BytesIO(corrupt_zip), 'x.zip'),
                                      'csrf_token': SUITE_CSRF},
                     content_type='multipart/form-data', follow_redirects=True)
        page = r.get_data(as_text=True)
        check(r.status_code == 200 and 'could not be read and were skipped' in page
              and 'PROJ0002' in page,
              'the upload page tells the user which pair was skipped and why',
              ' '.join(page[page.find('skipped:'):page.find('skipped:') + 100].split())
              if 'skipped:' in page else f'status {r.status_code}')

        # ── peer-push failure is recorded, not printed ──
        import urllib.error as _ue
        pc = sys.modules['peer_client']
        pc.PEER_URL = 'http://peer.invalid'
        pc.IMPORT_KEY = 'k'
        _real_urlopen = pc.urllib.request.urlopen

        def _down(req, timeout=None):
            raise _ue.URLError('connection refused (test)')
        pc.urllib.request.urlopen = _down
        try:
            t = pc.push_to_peer(full)
            check(t is not None, 'push_to_peer returns its thread')
            t.join(10)
            pe = pi.sync.get_last_push_error()
            check(pe is not None and 'connection refused' in pe['error'],
                  'a failed push is recorded in sync_state', str(pe))
            check(pe and pe.get('at') and pe.get('dates') == SESSION_DATE,
                  'with a timestamp and the dates it carried', str(pe))
            r = _c2.get('/upload')
            check('last push to the other Pi failed' in r.get_data(as_text=True),
                  'the upload page shows the failure')

            class _Resp:
                def __enter__(self): return self
                def __exit__(self, *a): return False
                def read(self): return b'{"imported":1}'

            pc.urllib.request.urlopen = lambda req, timeout=None: _Resp()
            t = pc.push_to_peer(full)
            t.join(10)
            check(pi.sync.get_last_push_error() is None, 'a later success clears it')
        finally:
            pc.urllib.request.urlopen = _real_urlopen
            pc.PEER_URL = ''
        pi.assess.delete_assessment(aid_f)

        # ── 5. deletions replicate ────────────────────────────────────────────
        print('\n5. session deletions replicate to the peer')
        baid = b.assess.create_assessment('Peer side', standard='bs4142')
        bloc = b.assess.add_assessment_location(baid, 'Loc B')
        b.assess.assign_runs(baid, bloc, [(SESSION_DATE, 9)])
        check(n_of(b, 'assessment_runs') == 1, 'peer has an assigned run')

        # 5a — a local delete leaves a tombstone and publishes it
        a.db.delete_session(SESSION_DATE)
        tomb = a.tombstones()
        check(SESSION_DATE in tomb, 'delete_session records a tombstone',
              f'deleted_at={tomb.get(SESSION_DATE)}')
        full = a.sync.get_full_sync_payload()
        check([t['date'] for t in full['deleted_sessions']] == [SESSION_DATE],
              'tombstone is carried in the full sync payload')

        # 5b — a peer that was offline catches up on the next full sync
        check(b.has_session(SESSION_DATE), 'peer still has the session beforehand')
        b.set_imported_at(SESSION_DATE, '2020-01-01 00:00:00')   # imported long ago
        b.sync.apply_full_sync(full)
        check(not b.has_session(SESSION_DATE),
              'offline peer deletes the session on full-sync catch-up')
        check(n_of(b, 'runs') == 0, 'peer runs cascaded')
        check(n_of(b, 'assessment_runs') == 0, 'peer assessment_runs cleared')
        # Pre-WP10 this asserted == 1: the sender's assessment (same local id)
        # silently overwrote the peer's own on every full sync — the F6
        # finding. uid-keyed replication keeps the peer's row AND lands the
        # sender's as rows of their own.
        check(n_of(b, 'assessments') == n_of(a, 'assessments') + 1,
              "peer's own assessment survives, and the sender's arrive beside it "
              '(uid-keyed since WP10)',
              f"b={n_of(b, 'assessments')}, a={n_of(a, 'assessments')}")
        check(b.sql("SELECT COUNT(*) FROM assessments WHERE name='Peer side'")[0][0] == 1,
              "and the peer's row was not renamed by an id collision")
        check(b.tombstones().get(SESSION_DATE) == tomb[SESSION_DATE],
              'peer relays the tombstone with the original deletion time')

        # 5c — re-importing a date clears its tombstone
        b.db.import_sessions(payload)
        check(b.has_session(SESSION_DATE), 'peer re-imports the date')
        check(SESSION_DATE not in b.tombstones(),
              're-import clears the tombstone, so it will not be re-deleted')

        # 5d — a stale tombstone must not delete a newer local re-import
        b.set_imported_at(SESSION_DATE, '2030-01-01 00:00:00')   # newer than the delete
        b.sync.apply_full_sync(full)
        check(b.has_session(SESSION_DATE),
              'a re-import newer than the peer tombstone is NOT deleted')

        # 5e — the live event path deletes and re-tombstones
        b.set_imported_at(SESSION_DATE, '2020-01-01 00:00:00')
        b.sync.apply_sync_event('session', 'delete', {'date': SESSION_DATE})
        check(not b.has_session(SESSION_DATE), "apply_sync_event('session','delete') deletes")
        check(SESSION_DATE in b.tombstones(),
              'the replayed delete tombstones locally too, so it keeps propagating')

        # 5f — bulk purge replicates the same way
        print('\n6. purge_sessions_before() replicates')
        for side in (a, b):
            side.exec("INSERT OR REPLACE INTO sessions (date, run_count, avg_laeq, max_laeq) "
                      "VALUES ('2026-07-01', 1, 50.0, 60.0)")
            side.exec("INSERT OR REPLACE INTO sessions (date, run_count, avg_laeq, max_laeq) "
                      "VALUES ('2026-07-15', 1, 51.0, 61.0)")
            side.exec("INSERT OR REPLACE INTO sessions (date, run_count, avg_laeq, max_laeq) "
                      "VALUES ('2026-09-01', 1, 52.0, 62.0)")
            side.exec("UPDATE sessions SET imported_at='2020-01-01 00:00:00'")
        check(n_of(b, 'sessions') == 3, 'peer has three sessions before the purge')

        purged = a.db.purge_sessions_before('2026-08-01')
        check(sorted(purged) == ['2026-07-01', '2026-07-15'],
              'purge removes only dates before the cutoff', str(purged))
        atomb = a.tombstones()
        check(all(d in atomb for d in purged), 'purge tombstones every date it removed')
        check('2026-09-01' not in atomb, 'the surviving date is not tombstoned')

        b.sync.apply_sync_event('session', 'purge_before', {'before': '2026-08-01'})
        check(not b.has_session('2026-07-01') and not b.has_session('2026-07-15'),
              'purge event removes both dates on the peer')
        check(b.has_session('2026-09-01'), 'purge event leaves later sessions alone')
        check(all(d in b.tombstones() for d in purged),
              'peer tombstones the purged dates too')

        # and a peer that missed the event still catches up via full sync
        c = Side(os.path.join(tmp, 'c.db'))
        c.db.import_sessions(payload)
        c.exec("INSERT OR REPLACE INTO sessions (date, run_count, avg_laeq, max_laeq) "
               "VALUES ('2026-07-01', 1, 50.0, 60.0)")
        c.exec("UPDATE sessions SET imported_at='2020-01-01 00:00:00'")
        check(c.has_session('2026-07-01') and c.has_session(SESSION_DATE),
              'third Pi starts with both dates')
        c.sync.apply_full_sync(a.sync.get_full_sync_payload())
        check(not c.has_session('2026-07-01'), 'purged date removed on catch-up')
        check(not c.has_session(SESSION_DATE), 'deleted session removed on catch-up')

        # ── 6b. session aggregates and orphan cleanup ─────────────────────────
        print('\n6b. recompute_session_aggregates() and delete_orphaned_runs()')
        m = Side(os.path.join(tmp, 'maint.db'))
        m.db.import_sessions(sessions, metadata=META)
        good = m.sql('SELECT avg_laeq, max_laeq, run_count FROM sessions WHERE date=?',
                     SESSION_DATE)[0]
        import_avg = good['avg_laeq']

        # Reproduce the real defect: runs corrected, session left stale (the old
        # 0x0422 LAeq bug survived at session level ~8 dB high).
        m.exec('UPDATE sessions SET avg_laeq=83.58, max_laeq=1.0, run_count=99 WHERE date=?',
               SESSION_DATE)
        changed = m.db.recompute_session_aggregates([SESSION_DATE])
        fixed = m.sql('SELECT avg_laeq, max_laeq, run_count FROM sessions WHERE date=?',
                      SESSION_DATE)[0]
        check(len(changed) == 1 and changed[0][0] == SESSION_DATE,
              'recompute reports the changed session', str(changed))
        check(fixed['avg_laeq'] == import_avg,
              'recomputed avg_laeq equals what import produced',
              f'{fixed["avg_laeq"]} vs {import_avg}')
        check(fixed['max_laeq'] == good['max_laeq'], 'max_laeq restored')
        check(fixed['run_count'] == good['run_count'], 'run_count restored')
        check(m.db.recompute_session_aggregates([SESSION_DATE]) == [],
              'recompute is idempotent — second pass changes nothing')

        # Orphans: a run pointing at a session id that does not exist, plus one
        # with a NULL session_id, must both go; attached runs must survive.
        attached_before = n_of(m, 'runs', 'WHERE session_id IN (SELECT id FROM sessions)')
        m.exec_nofk('INSERT INTO runs (session_id, run_number, avg_laeq) VALUES (999999, 1, 50.0)')
        m.exec_nofk('INSERT INTO runs (session_id, run_number, avg_laeq) VALUES (NULL, 2, 51.0)')
        check(n_of(m, 'runs') == attached_before + 2, 'two orphan rows planted')
        removed = m.db.delete_orphaned_runs()
        check(removed == 2, 'delete_orphaned_runs removed exactly the orphans', str(removed))
        check(n_of(m, 'runs') == attached_before, 'attached runs untouched',
              str(attached_before))
        check(m.db.delete_orphaned_runs() == 0, 'idempotent — nothing left to remove')

        # ── 6c. report template integrity ─────────────────────────────────────
        print('\n6c. seed report templates are structurally valid')
        SECTIONS = ('executive_summary', 'methodology', 'results_narrative',
                    'compliance', 'conclusions', 'recommendations')
        for t in m.reports_db.DEFAULT_TEMPLATES:
            nm = t['name']
            # Without this token the prompt reaches Claude with no measurement
            # data in it and the report is confidently invented.
            check('{{session_data}}' in t['prompt'],
                  f'{nm}: has the {{{{session_data}}}} token')
            missing = [k for k in SECTIONS if k not in t['prompt']]
            check(not missing, f'{nm}: names all six report sections', str(missing))
        check(sum(t['is_default'] for t in m.reports_db.DEFAULT_TEMPLATES) == 1,
              'exactly one seed template is the default')
        # BS 8233:2014 Table 4 — bedrooms are 30 dB LAeq,8h at night, not 35.
        plan = next(t for t in m.reports_db.DEFAULT_TEMPLATES
                    if t['name'] == 'Planning Noise Assessment')
        check('30 dB LAeq,8h at night' in plan['prompt'],
              'Planning: bedroom night guideline is 30 dB LAeq,8h')
        check('bedrooms: 35 dB LAeq,8h night' not in plan['prompt'],
              'Planning: the old incorrect bedroom value is gone')
        check('rating level' in plan['prompt'],
              'Planning: BS 4142 rating level (not raw specific level) is required')

        # The Noise Act permitted level depends on the underlying sound level and is
        # assessed indoors; there is no single number to compare an external LAeq to.
        lge = next(t for t in m.reports_db.DEFAULT_TEMPLATES
                   if t['name'] == 'Local Government Enforcement')
        check('35 dB LAeq indoors' not in lge['prompt'],
              'LGE: the old fixed 35/45 dB Noise Act figures are gone')
        check('underlying level' in lge['prompt'] and '34 dB LAeq' in lge['prompt'],
              'LGE: permitted level stated as depending on the underlying level')
        check('dwelling' in lge['prompt'], 'LGE: says it is assessed inside the dwelling')

        # Workplace action values are a daily personal exposure, not a measured LAeq.
        occ = next(t for t in m.reports_db.DEFAULT_TEMPLATES
                   if t['name'] == 'Occupational Health')
        check('LEP,d' in occ['prompt'], 'Occupational: uses LEP,d')
        check('dB LAeq,8h /' not in occ['prompt'],
              'Occupational: no longer calls the action values LAeq,8h')

        # ── 7. NOR140 xlsx parity with every Nortfr reference present ─────────
        print('\n7. xlsx export matches the Nortfr references')
        check_references(meas_root, tmp)

        # ── 8. WP4/F9 — CSV headers name the channel actually written ─────────
        # Reuses the Flask test client and 'rt' Side set up in section 4d — its
        # DB has the real SD-card sessions imported and noise_app is already
        # bound to it (the first import of noise_app in this process binds its
        # `from noise_db import ...` names to whichever noise_db module is
        # current at that moment, which was rt's).
        print('\n8. CSV export headers match their contents (F9)')

        resp = _client.get('/export/sessions.csv')
        check(resp.status_code == 200, 'sessions CSV downloads', resp.status_code)
        sess_header = resp.get_data(as_text=True).split('\r\n')[0].split(',')
        check('lafmax_max_db' in sess_header,
              'sessions CSV: lafmax_max_db header present (was laeq_max_db, held LAFmax)',
              sess_header)
        check('laeq_max_db' not in sess_header,
              'sessions CSV: stale laeq_max_db header gone', sess_header)
        check('laeq_avg_db' in sess_header,
              'sessions CSV: laeq_avg_db kept (WP5 changes its definition, not its name)',
              sess_header)

        aid8 = rt.assess.create_assessment('F9 CSV test')
        lid8 = rt.assess.add_assessment_location(aid8, 'L')
        rt.assess.assign_runs(aid8, lid8, [(SESSION_DATE, 9)])
        resp = _client.get(f'/export/assessment/{aid8}.csv')
        check(resp.status_code == 200, 'assessment CSV downloads', resp.status_code)
        lines = resp.get_data(as_text=True).split('\r\n')
        assess_header = lines[0].split(',')
        check('lafmin_db' in assess_header and 'lafmax_db' in assess_header,
              'assessment CSV: lafmin_db/lafmax_db headers present '
              '(were min_laeq_db/max_laeq_db, held LAFmin/LAFmax)', assess_header)
        check('min_laeq_db' not in assess_header and 'max_laeq_db' not in assess_header,
              'assessment CSV: stale min/max_laeq_db headers gone', assess_header)
        check('n_samples' in assess_header,
              'assessment CSV: n_samples header present (was mislabelled duration_s)',
              assess_header)
        check('duration_s' in assess_header,
              'assessment CSV: duration_s is now the real meter-stored value', assess_header)
        row1 = lines[1].split(',')
        check(len(row1) == len(assess_header),
              'assessment CSV: data row width matches header',
              f'{len(row1)} vs {len(assess_header)}')
        rt.assess.delete_assessment(aid8)

        # Per-run CSV headers are built client-side in index.html — assert the
        # source names the real fields (grep, since it's JS, not testable via
        # the Python data layer).
        idx_html = open(os.path.join(REPO, 'templates', 'index.html'), encoding='utf-8').read()
        run_hdr_line = [l for l in idx_html.split('\n') if "'run', 'start_time'" in l]
        check(len(run_hdr_line) == 1, 'exportSessionRuns header row found once',
              str(len(run_hdr_line)))
        hl = run_hdr_line[0]
        check('lafmin_db' in hl and 'lafmax_db' in hl,
              'exportSessionRuns: lafmin_db/lafmax_db headers present '
              '(were laeq_min_db/laeq_max_db, held LAFmin/LAFmax)', hl.strip())
        check('laeq_min_db' not in hl and 'laeq_max_db' not in hl,
              'exportSessionRuns: stale laeq_min_db/laeq_max_db gone', hl.strip())
        check("'duration_s'" in hl,
              'exportSessionRuns: duration_s header present (was p.n, the record count)',
              hl.strip())

        run_csv_fn = idx_html.split('function exportRunCSV(')[1].split('\nfunction ')[0]
        check("'lae_db'" in run_csv_fn,
              'exportRunCSV: lae_db header present (was laimax_db, holds the LAE channel)',
              run_csv_fn[:300])
        check("'lapeak_db'" in run_csv_fn,
              'exportRunCSV: lapeak_db header present (was lcpeak_db, holds LApeak, not true LCpeak)')
        check("'laimax_db'" not in run_csv_fn, 'exportRunCSV: stale laimax_db header gone')
        check("'lcpeak_db'" not in run_csv_fn, 'exportRunCSV: stale lcpeak_db header gone')

        # ── 9. security: startup, CSRF, API key, rate limiting ────────────────
        # Everything here drives the real app through the Flask test client.
        # `rt` (section 4d) is the Side noise_app is bound to: the module was
        # imported while that binding was current and captured its functions by
        # value, so the app reads and writes route.db no matter which Side was
        # constructed since.
        import logging as _lg
        import subprocess as _sp
        _wa = importlib.import_module('webauth')
        _napp = importlib.import_module('noise_app')
        _sapp = _napp.app

        print('\n9a. the app refuses to start with authentication missing')
        # The Mac has no flight-tracker auth module, which is exactly the
        # condition that used to make every page public.
        check(_wa.AUTH_AVAILABLE is False,
              'the auth module is absent here, as on any dev machine')
        _saved_allow = os.environ.pop('ALLOW_UNAUTHENTICATED', None)
        try:
            try:
                _wa.check_startup_security()
                _refused = False
            except _wa.InsecureConfiguration:
                _refused = True
            check(_refused, 'check_startup_security() refuses without the override')
        finally:
            if _saved_allow is not None:
                os.environ['ALLOW_UNAUTHENTICATED'] = _saved_allow
        _wa.check_startup_security()   # with the override set: must not raise
        check(True, 'and starts when ALLOW_UNAUTHENTICATED=1 is set')

        # The decisive check: a real interpreter importing the real module.
        # A refusal that only exists in a helper nobody calls is not a refusal.
        _env = dict(os.environ)
        _env.pop('ALLOW_UNAUTHENTICATED', None)
        _env['NOISE_DB_PATH'] = os.path.join(tmp, 'refuse.db')
        _proc = _sp.run([sys.executable, '-c', 'import noise_app'], cwd=REPO,
                        env=_env, capture_output=True, text=True)
        check(_proc.returncode != 0, 'importing noise_app without the override exits non-zero',
              f'rc={_proc.returncode}')
        check('InsecureConfiguration' in _proc.stderr,
              'and says why', _proc.stderr.strip().split('\n')[-1][:100])

        # An empty UPLOAD_PASSWORD leaves the upload form open; that must be
        # said out loud at startup rather than discovered later.
        _seen = []

        class _Capture(_lg.Handler):
            def emit(self, record):
                _seen.append(record)

        _cap = _Capture()
        _lg.getLogger('noise.webauth').addHandler(_cap)
        try:
            _wa.check_startup_security()
        finally:
            _lg.getLogger('noise.webauth').removeHandler(_cap)
        check(_wa.UPLOAD_PASS == '', 'no upload password is set in this environment')
        check(any(r.levelno == _lg.WARNING and 'UPLOAD_PASSWORD' in r.getMessage()
                  for r in _seen), 'an empty UPLOAD_PASSWORD is warned about at startup')
        check(any(r.levelno == _lg.ERROR for r in _seen),
              'and running unauthenticated is logged at ERROR')

        print('\n9b. session cookie flags')
        check(_sapp.config['SESSION_COOKIE_SAMESITE'] == 'Lax', 'SameSite=Lax')
        check(_sapp.config['SESSION_COOKIE_HTTPONLY'] is True, 'HttpOnly')
        check(_sapp.config['SESSION_COOKIE_SECURE'] is True,
              'Secure (default on; SESSION_COOKIE_SECURE=0 for plain-http LAN use)')

        print('\n9c. CSRF on cookie-authenticated state changes')
        _anon = _sapp.test_client()
        with _anon.session_transaction() as _s:
            _s['user'] = 'test'
            _s['logged_in'] = True          # logged in, but no token
        _r = _anon.post(f'/session/{SESSION_DATE}/edit',
                        data={'recorder_name': 'Mallory'})
        check(_r.status_code == 403, 'a logged-in POST with no token is refused',
              str(_r.status_code))
        _name_now = rt.sql('SELECT recorder_name FROM sessions WHERE date=?',
                           SESSION_DATE)[0]['recorder_name']
        check(_name_now != 'Mallory', 'and the write did not land', repr(_name_now))

        _r = _anon.post(f'/session/{SESSION_DATE}/edit',
                        data={'recorder_name': 'Mallory', 'csrf_token': 'not-the-token'})
        check(_r.status_code == 403, 'a wrong token is refused too', str(_r.status_code))

        _good = _sapp.test_client()
        with _good.session_transaction() as _s:
            _s['user'] = 'test'
            _s['logged_in'] = True
            _s['_csrf_token'] = SUITE_CSRF
        _r = _good.post(f'/session/{SESSION_DATE}/edit', headers=dict(
            CSRF_HDR, **{'X-Requested-With': 'XMLHttpRequest'}),
            data={'recorder_name': 'Catherine Ives-Yim'})
        check(_r.status_code == 200, 'the token in the X-CSRF-Token header is accepted',
              f'{_r.status_code}: {_r.get_data(as_text=True)[:80]}')
        _r = _good.post(f'/session/{SESSION_DATE}/edit',
                        data={'recorder_name': 'Catherine Ives-Yim',
                              'csrf_token': SUITE_CSRF})
        check(_r.status_code != 403, 'the token in a form field is accepted too',
              str(_r.status_code))
        check(rt.sql('SELECT recorder_name FROM sessions WHERE date=?',
                     SESSION_DATE)[0]['recorder_name'] == 'Catherine Ives-Yim',
              'and that write did land')
        check(_anon.get('/health').status_code == 200, 'GET is not affected')

        print('\n9d. the API key: header or form, never the query string')
        _saved_key = _wa.IMPORT_KEY
        _wa.IMPORT_KEY = 'a-long-random-import-key'
        try:
            _r = _anon.get(f'/api/sync?api_key={_wa.IMPORT_KEY}')
            check(_r.status_code == 403, 'a key in the query string is not accepted',
                  str(_r.status_code))
            _r = _anon.get('/api/sync', headers={'X-Import-Key': _wa.IMPORT_KEY})
            check(_r.status_code == 200, 'the same key in the header is',
                  str(_r.status_code))
            _r = _anon.get('/api/sync', headers={'X-Import-Key': 'a-long-random-import-kez'})
            check(_r.status_code == 403, 'a near-miss key is refused', str(_r.status_code))
            # An API-key request carries no cookie to ride on, so it is CSRF
            # exempt — import_sdcard.py and the peer Pi never see a page.
            _r = _anon.post('/import', headers={'X-Import-Key': _wa.IMPORT_KEY},
                            json={'sessions': []})
            check(_r.status_code == 400 and 'no sessions' in _r.get_data(as_text=True),
                  'an API-key POST needs no CSRF token', str(_r.status_code))
            _r = _anon.post('/import', json={'sessions': []})
            check(_r.status_code == 403, 'and without the key it is still refused',
                  str(_r.status_code))
        finally:
            _wa.IMPORT_KEY = _saved_key

        print('\n9e. the templates carry the token (the client half is not Python)')
        _tdir = os.path.join(REPO, 'templates')

        def _tpl_text(name):
            return open(os.path.join(_tdir, name), encoding='utf-8').read()

        for _name in ('index.html', 'manage.html', 'upload.html',
                      'assessments.html', 'reports.html', 'report.html', 'login.html'):
            check("{% include '_csrf.html' %}" in _tpl_text(_name),
                  f'{_name}: includes the CSRF partial')
        _partial = _tpl_text('_csrf.html')
        check("name=\"csrf-token\"" in _partial and 'X-CSRF-Token' in _partial,
              'the partial publishes the token and sets the fetch header')
        for _name, _n in (('manage.html', 3), ('login.html', 3), ('upload.html', 1)):
            _txt = _tpl_text(_name)
            _forms = _txt.lower().count('<form ')
            _fields = _txt.count('name="csrf_token"')
            check(_fields == _n == _forms,
                  f'{_name}: every posting form has a hidden csrf_token field',
                  f'{_fields} fields / {_forms} forms')
        for _name in os.listdir(_tdir):
            if _name.endswith('.html'):
                check('?api_key=' not in _tpl_text(_name),
                      f'{_name}: no API key in a query string')

        # A template that fails to render is a 500 on a page nothing else in
        # the suite visits, so every page is fetched and the token looked for.
        for _path in ('/', '/manage', '/upload', '/login', '/assessments', '/reports'):
            _r = _good.get(_path)
            check(_r.status_code == 200, f'GET {_path} renders', str(_r.status_code))
            check('name="csrf-token"' in _r.get_data(as_text=True),
                  f'GET {_path} publishes the token')

        print('\n9f. rate limiting on /login')
        if _napp.limiter is None:
            print('  --    Flask-Limiter not installed in this venv; limit not exercised')
        else:
            _lim = _sapp.test_client()
            with _lim.session_transaction() as _s:
                _s['_csrf_token'] = SUITE_CSRF
            _codes = [_lim.post('/login', headers=CSRF_HDR,
                                data={'step': 'email', 'email': f'a{i}@example.com',
                                      'csrf_token': SUITE_CSRF}).status_code
                      for i in range(11)]
            check(_codes[:10] == [200] * 10, 'ten login attempts a minute are allowed',
                  str(_codes[:10]))
            check(_codes[10] == 429, 'the eleventh is rate limited', str(_codes[10]))
            check(_lim.get('/login').status_code == 200,
                  'but the login page itself is still servable — the limit is POST only')

        # ── 10. the run modal names the quantity it shows (F7) ─────────────────
        print('\n10. run modal quantities and the 1-second series route')
        # These numbers are what an officer reads off the screen and puts in a
        # statement, and three of them named the wrong measurement: the modal
        # labelled the downsampled LApeak channel "LCpeak max", the LAE channel
        # "LAImax peak", and the extremes of a max-of-window profile
        # "LAmax"/"LAmin". None of it was reachable from Python — the display is
        # JavaScript — so the guard below is static, backed by a live check of
        # the route the modal now fetches its series from.
        #
        # Its own Side, and noise_app dropped with it: the module binds
        # noise_db's functions at import, so a stale noise_app answers from
        # whichever database it was first imported against.
        sys.modules.pop('noise_app', None)
        pf = Side(os.path.join(tmp, 'prof.db'))
        pf.db.import_sessions(sessions, metadata=META)
        _papp = importlib.import_module('noise_app').app
        _papp.config['TESTING'] = True
        _pcl = _papp.test_client()
        with _pcl.session_transaction() as _s:
            _s['user'] = 'test'
            _s['logged_in'] = True

        # A downsampled run: the chart profile is one point per `step` seconds,
        # each of them the maximum of its window. Take the one where that hurts
        # most, so the checks below are made against a real disagreement rather
        # than a run that happens never to approach 85 dB.
        _pct = lambda xs: 100.0 * sum(1 for v in xs if v >= 85) / len(xs)
        _expand = lambda p, n: [v for v in p['laeq_profile'] for _ in range(p['step'])][:n]
        _worst, _over = None, -1.0
        for _p in pf.db.get_all_sessions_json()['sessions'][0]['projects']:
            _full = pf.db.get_run_prof_by_source(SESSION_DATE, _p['source_file'])['prof_laeq']
            if not _full or _p['step'] <= 1:
                continue
            _d = _pct(_expand(_p, len(_full))) - _pct(_full)
            if _d > _over:
                _worst, _over = _p, _d
        check(_over > 1.0, 'a run whose chart profile materially overstates time above 85 dB',
              f"{_worst['source_file']}: +{_over:.1f} percentage points")
        _dsrow = pf.sql('SELECT r.source_file, r.step, r.n_samples, r.prof_laeq_json '
                        'FROM runs r JOIN sessions s ON s.id=r.session_id '
                        'WHERE s.date=? AND r.source_file=?', SESSION_DATE, _worst['source_file'])[0]
        _src8 = _dsrow['source_file']
        _stored = json.loads(_dsrow['prof_laeq_json'])

        _r = _pcl.get(f'/api/run/{SESSION_DATE}/{_src8}/prof')
        check(_r.status_code == 200, 'the prof route serves a run that exists',
              f'{_r.status_code}: {_r.get_data(as_text=True)[:120]}')
        _j = _r.get_json()
        check(_j['prof_laeq'] == _stored,
              'it returns the stored 1-second series unaltered', f"{len(_j['prof_laeq'])} values")
        check(len(_j['prof_laeq']) == _j['n'] == _dsrow['n_samples'],
              'one value per second of the run', str(_j['n']))

        # The difference the fix is for, on real data: counting the expanded
        # chart profile against 85 dB is not counting the measurement.
        _proj8 = next(p for p in pf.db.get_all_sessions_json()['sessions'][0]['projects']
                      if p['source_file'] == _src8)
        _chart = _expand(_proj8, _j['n'])
        check(len(_proj8['laeq_profile']) < len(_j['prof_laeq']),
              'the browser payload really does carry only the downsampled profile',
              f"{len(_proj8['laeq_profile'])} vs {len(_j['prof_laeq'])}")
        check(_pct(_chart) > _pct(_stored),
              'counting the chart profile overstates the time above 85 dB',
              f'chart {_pct(_chart):.1f}% vs real {_pct(_stored):.1f}%')
        check(min(_proj8['laeq_profile']) > min(_stored),
              'and its minimum is biased high — so it cannot be an "LAmin"',
              f'{min(_proj8["laeq_profile"])} vs {min(_stored)}')

        # Keyed on the stable identity, so a renumbering between render and
        # fetch cannot swap the measurement under an open modal.
        _sid8 = pf.sql('SELECT id FROM sessions WHERE date=?', SESSION_DATE)[0]['id']
        pf.exec('UPDATE runs SET run_number=run_number+100 WHERE session_id=?', _sid8)
        _r = _pcl.get(f'/api/run/{SESSION_DATE}/{_src8}/prof')
        check(_r.status_code == 200 and _r.get_json()['prof_laeq'] == _stored,
              'a renumbered run still resolves by source_file', str(_r.status_code))
        pf.exec('UPDATE runs SET run_number=run_number-100 WHERE session_id=?', _sid8)

        # A run whose PROF series was never stored — an older peer-synced
        # session — answers 200 with an empty series, which is what tells the
        # page to fall back to the chart profile and label it, rather than
        # showing an empty distribution as if the run had been silent.
        pf.exec('UPDATE runs SET prof_laeq_json=NULL WHERE session_id=? AND source_file=?',
                _sid8, _src8)
        _r = _pcl.get(f'/api/run/{SESSION_DATE}/{_src8}/prof')
        check(_r.status_code == 200 and _r.get_json()['prof_laeq'] == [],
              'a run with no stored series is 200 with an empty series, not a 404',
              str(_r.status_code))
        pf.exec('UPDATE runs SET prof_laeq_json=? WHERE session_id=? AND source_file=?',
                json.dumps(_stored), _sid8, _src8)

        # An unknown run is a 404, not an empty series presented as a result.
        _r = _pcl.get(f'/api/run/{SESSION_DATE}/PROJ7777/prof')
        check(_r.status_code == 404, 'an unknown source_file is a 404', str(_r.status_code))
        _r = _pcl.get(f'/api/run/1999-01-01/{_src8}/prof')
        check(_r.status_code == 404, 'an unknown date is a 404', str(_r.status_code))
        check(pf.db.get_run_prof_by_source(SESSION_DATE, 'PROJ7777') is None,
              'the helper returns None rather than an empty series')

        # The page still renders — a template that fails to parse takes the
        # whole browser view with it.
        check(_pcl.get('/').status_code == 200, 'the session browser still renders')

        # ── 8b. static guard on the modal label -> field mapping ──────────────
        print('\n10b. every modal row is bound to the field its label names')
        _idx = open(os.path.join(REPO, 'templates', 'index.html'), encoding='utf-8').read()

        def _jsfn(name):
            """Source of one top-level function in index.html."""
            _ls = _idx.split('\n')
            _st = next(i for i, l in enumerate(_ls)
                       if l.startswith(f'function {name}(') or l.startswith(f'async function {name}('))
            _en = next(i for i in range(_st + 1, len(_ls)) if _ls[i] == '}')
            return '\n'.join(_ls[_st:_en + 1])

        _render = _jsfn('renderModalStats')
        _openfn = _jsfn('openRunModal')
        _loadfn = _jsfn('loadRunProf')

        def _row(label):
            hits = [l.strip() for l in _render.split('\n') if label in l]
            check(len(hits) == 1, f'exactly one modal row for {label}', str(len(hits)))
            return hits[0]

        # The displayed value, not merely a mention of the field somewhere on the
        # line: a row guarded by `proj.pmx != null` while showing
        # max(lcpeak_profile) is exactly the defect, and naming the whole line
        # would have let it through.
        check('d1(proj.pmx)' in _row('LC<sub>peak</sub> max'),
              'LCpeak max displays the GLOB LCpeak scalar (proj.pmx)')
        check('d1(proj.laimax)' in _row('LA<sub>Imax</sub>'),
              'LAImax displays the GLOB LAImax scalar')
        check('d1(proj.mx)' in _row('LA<sub>Fmax</sub>'), 'LAFmax displays proj.mx')
        check('d1(proj.mn)' in _row('LA<sub>Fmin</sub>'), 'LAFmin displays proj.mn')
        check('d1(proj.avg)' in _row('LA<sub>eq</sub> (energy avg)'),
              'LAeq is the stored GLOB LAeq, not a profile energy average')
        check("'LA<sub>max</sub>'" not in _idx and "'LA<sub>min</sub>'" not in _idx,
              'the ambiguous LAmax/LAmin rows are gone')

        # And no row may come off a downsampled profile at all — the general
        # form of the same mistake, whichever channel it is made with next.
        _rowlines = [l.strip() for l in _render.split('\n')
                     if "'dB'," in l or "'%'," in l]
        check(len(_rowlines) >= 9, 'the guard located the modal rows', str(len(_rowlines)))
        _fromprof = [l for l in _rowlines if '_profile' in l]
        check(not _fromprof, 'no modal row is derived from a downsampled profile',
              str(_fromprof))

        check('computeStats(prof.series)' in _render,
              'the series-derived rows are computed from the fetched series')
        check('pct85' in _row('Time ≥ 85 dB') and 'stats ? stats.pct85' in _render,
              'Time ≥ 85 dB comes from that series')
        check("drawHistogram(document.getElementById('modalHistCanvas'), prof.series)" in _openfn,
              'the histogram bins the fetched series')
        check("drawExceedance(document.getElementById('modalExcCanvas'), prof.series)" in _openfn,
              'the exceedance curve reads the fetched series')
        check('laeq_profile' not in _openfn,
              'nothing the modal counts is taken from the chart profile')

        check('${encodeURIComponent(proj.source_file)}/prof' in _loadfn,
              'the modal fetches the series by source_file, not by run number')
        check("source: 'chart'" in _loadfn,
              'a run with no stored series falls back to the chart profile')
        check('chart profile' in _render,
              'and the fallback says so on the rows it produced')

        # ── 11. report provenance: pinned runs and input snapshots (F4) ───────
        print('\n11. report provenance: pinned runs and input snapshots (F4)')
        # generated_reports stored only run_number, and view_report selected
        # all_projects[run_number-1] — so once a re-import renumbered the
        # session, a stored report re-pointed at a different physical run. And
        # the statistics table was recomputed live while the narrative was
        # frozen, so a later backfill made a report disagree with itself. The
        # report now pins its run by source_file and renders from a snapshot
        # of its inputs; live values only ever produce a notice.
        #
        # Its own Side, noise_app dropped with it (see section 10). Claude is
        # stubbed at the module boundary: the route reads _call_claude through
        # the reports module noise_app registered, which is this Side's.
        sys.modules.pop('noise_app', None)
        rp = Side(os.path.join(tmp, 'prov.db'))
        rp.db.import_sessions(_copy.deepcopy(sessions), metadata=META)
        _rapp = importlib.import_module('noise_app').app
        _rapp.config['TESTING'] = True
        _rcl = _rapp.test_client()
        with _rcl.session_transaction() as _s:
            _s['user'] = 'test'
            _s['logged_in'] = True
            _s['_csrf_token'] = SUITE_CSRF
        _FAKE_SECTIONS = {k: f'<p>{k}</p>' for k in (
            'executive_summary', 'methodology', 'results_narrative',
            'compliance', 'conclusions', 'recommendations')}
        _prompts = []

        def _fake_claude(prompt, model, thinking_level):
            _prompts.append(prompt)
            return _FAKE_SECTIONS, 100, 50, 0.0015
        rp.reports._call_claude = _fake_claude

        def _gr(rid):
            return dict(rp.sql('SELECT * FROM generated_reports WHERE id=?', rid)[0])

        def _src_of(n):
            return rp.sql('SELECT r.source_file FROM runs r JOIN sessions s ON s.id=r.session_id '
                          'WHERE s.date=? AND r.run_number=?', SESSION_DATE, n)[0]['source_file']

        check(_src_of(5) == 'PROJ0005', 'baseline: run 5 is PROJ0005', _src_of(5))

        # (a) Generated with the run's identity, as both clients now send it.
        _r = _rcl.post('/api/generate-report', headers=CSRF_HDR, json={
            'session_date': SESSION_DATE, 'run_number': 5, 'source_file': 'PROJ0005',
            'model': 'claude-sonnet-5', 'thinking_level': 'none'})
        check(_r.status_code == 200, 'generate-report for run 5 returns 200',
              f'{_r.status_code} {_r.get_data(as_text=True)[:200]}')
        rid5 = _r.get_json()['report_id']
        row5 = _gr(rid5)
        check(row5['source_file'] == 'PROJ0005', 'the report stores the run source_file',
              str(row5['source_file']))
        check(row5['run_number'] == 5 and row5['run_label'] == 'Run 5',
              'and the run number as a label', f"{row5['run_number']} {row5['run_label']}")
        snap5 = json.loads(row5['input_snapshot_json'] or 'null')
        check(isinstance(snap5, dict), 'an input snapshot is stored')
        check(set(snap5) >= {'session_data_block', 'stats', 'run_rows', 'sess',
                             'total_s', 'generated_from'},
              'the snapshot carries every input view_report renders', str(sorted(snap5)))
        check(snap5['run_rows'][0]['source_file'] == 'PROJ0005' and snap5['run_rows'][0]['run'] == 5,
              'snapshot run row names the run by source_file and number')
        # WP8: generated_from also names the instrument (the default serial,
        # '' on this Side, since the request named none).
        check(snap5['generated_from'] == {'dates': [SESSION_DATE], 'serials': [''],
                                          'source_files': ['PROJ0005']},
              'generated_from records the date, serial and source_file',
              str(snap5['generated_from']))
        check(snap5['session_data_block'] == _prompts[-1].split('\n\n', 1)[1]
              .rsplit('\n\nProduce a professional', 1)[0]
              or snap5['session_data_block'] in _prompts[-1],
              'the stored data block is the one the prompt was built from')
        check(snap5['stats'].get('session_leq') is not None, 'snapshot stats hold the session LAeq')
        snap_leq5 = snap5['run_rows'][0]['leq']

        # (b) An older client that sends only run_number: resolved to the
        # source_file now, while the number is still current.
        _r = _rcl.post('/api/generate-report', headers=CSRF_HDR, json={
            'session_date': SESSION_DATE, 'run_number': 3})
        check(_r.status_code == 200, 'generate-report by run_number alone returns 200',
              str(_r.status_code))
        rid3 = _r.get_json()['report_id']
        check(_gr(rid3)['source_file'] == _src_of(3) == 'PROJ0003',
              'run_number alone is resolved to its source_file', str(_gr(rid3)['source_file']))

        # A whole-session report has no single run to pin.
        _r = _rcl.post('/api/generate-report', headers=CSRF_HDR, json={'session_date': SESSION_DATE})
        check(_r.status_code == 200, 'a whole-session report generates', str(_r.status_code))
        rid_all = _r.get_json()['report_id']
        snap_all = json.loads(_gr(rid_all)['input_snapshot_json'])
        check(_gr(rid_all)['source_file'] is None and
              len(snap_all['generated_from']['source_files']) == len(snap_all['run_rows']) > 1,
              'a whole-session report pins no single run but lists every source_file')

        # (c) Viewing it while nothing has changed: no notice, source shown.
        _h = _rcl.get(f'/reports/{rid5}').get_data(as_text=True)
        check('id="provenanceNotice"' not in _h, 'an unchanged report shows no provenance notice')
        check('PROJ0005' in _h, 'the report names its source run')
        check(f'<td>5</td>' in _h and str(snap_leq5) in _h, 'and shows run 5 and its LAeq')

        # (d) A WP1-style renumbering: PROJ0001 is gone from the card, so a
        # complete-date re-import deletes it and PROJ0005 becomes run 4.
        gone_p = _copy.deepcopy(sessions)
        gone_p[0]['projects'] = gone_p[0]['projects'][1:]
        gone_p[0]['complete_date'] = True
        gone_p[0]['skipped_files'] = []
        rp.db.import_sessions(gone_p, metadata=META)
        check(_src_of(4) == 'PROJ0005', 'after the re-import PROJ0005 is run 4', _src_of(4))
        _h = _rcl.get(f'/reports/{rid5}').get_data(as_text=True)
        check('data-state="differs"' in _h,
              'the report notices that live values differ from its snapshot')
        def _notice(html):
            i = html.find('<ul>', html.find('provenanceNotice'))
            return html[i:html.find('</ul>', i) + 5] if i > 0 else '(no notice)'
        check('PROJ0005 was run 5, now run 4' in _h, 'and says the run moved', _notice(_h))
        check(f'<td>5</td>' in _h and str(snap_leq5) in _h,
              'the table still shows the original run 5 and its figures')
        # The report generated by run_number alone is pinned the same way.
        check('PROJ0003 was run 3, now run 2' in _rcl.get(f'/reports/{rid3}').get_data(as_text=True),
              'a report generated by run_number alone follows its run too')

        # A backfill that changes a stored figure: the page keeps the snapshot
        # value and names the difference, rather than swapping silently.
        rp.exec('UPDATE runs SET avg_laeq = avg_laeq + 1.0 WHERE source_file=? AND session_id IN '
                '(SELECT id FROM sessions WHERE date=?)', 'PROJ0005', SESSION_DATE)
        _h = _rcl.get(f'/reports/{rid5}').get_data(as_text=True)
        _new_leq = round(snap_leq5 + 1.0, 1)
        check(f'PROJ0005 LAeq {snap_leq5} → {_new_leq}' in _h,
              'a changed stored LAeq is named in the notice', _notice(_h))
        check(f'<td class="">{snap_leq5}</td>' in _h or f'<td class="td-hi">{snap_leq5}</td>' in _h,
              'the table shows the snapshot LAeq, not the live one')
        check(f'<td class="">{_new_leq}</td>' not in _h and f'<td class="td-hi">{_new_leq}</td>' not in _h,
              'the live LAeq appears only in the notice')

        # (e) The run is gone altogether: the snapshot renders under a banner.
        rp.exec('DELETE FROM runs WHERE source_file=? AND session_id IN '
                '(SELECT id FROM sessions WHERE date=?)', 'PROJ0005', SESSION_DATE)
        _r = _rcl.get(f'/reports/{rid5}')
        check(_r.status_code == 200, 'a report whose run is gone still renders', str(_r.status_code))
        _h = _r.get_data(as_text=True)
        check('data-state="missing"' in _h and 'no longer stored' in _h,
              'with a banner saying the measurement is no longer stored')
        check(str(snap_leq5) in _h and '<td>5</td>' in _h, 'and the snapshot figures')

        # (f) A legacy row: no source_file, no snapshot. Renders live with a note.
        rp.exec('INSERT INTO generated_reports (session_date, run_number, run_label, model, '
                'thinking_level, sections_json) VALUES (?,?,?,?,?,?)',
                SESSION_DATE, 2, 'Run 2', 'claude-sonnet-5', 'none', json.dumps(_FAKE_SECTIONS))
        rid_legacy = rp.sql('SELECT MAX(id) AS id FROM generated_reports')[0]['id']
        _r = _rcl.get(f'/reports/{rid_legacy}')
        check(_r.status_code == 200, 'a legacy report renders', str(_r.status_code))
        _h = _r.get_data(as_text=True)
        check('data-state="no_snapshot"' in _h and 'before provenance was recorded' in _h,
              'with the no-snapshot note')
        _live2 = rp.reports._prepare_session_for_report(SESSION_DATE, run_number=2)[1][0]
        check(f'<td>2</td>' in _h and str(_live2['leq']) in _h and _live2['source_file'] in _h,
              'and the live figures for the run at that position')

        # (g) Listings carry the source_file and not the snapshot blob.
        _lst = _rcl.get('/api/generated-reports').get_json()
        _by_id = {r['id']: r for r in _lst}
        check(_by_id[rid5]['source_file'] == 'PROJ0005', 'api listing shows the source_file')
        check(all('input_snapshot_json' not in r for r in _lst),
              'and leaves the snapshot blob out')
        _h = _rcl.get('/reports').get_data(as_text=True)
        check('PROJ0005' in _h and 'PROJ0003' in _h, 'the history table names the pinned runs')
        check('data-src="' in _h, 'the run selector carries each run source_file')

        # (h) Migration: a database whose generated_reports predates the
        # provenance columns gets them, and single-run rows are backfilled
        # through the positional join — the one moment it is still trustworthy.
        _OLD_GR = """
            CREATE TABLE generated_reports (
              id INTEGER PRIMARY KEY, session_date TEXT NOT NULL, run_number INTEGER,
              run_label TEXT, template_id INTEGER, template_name TEXT, model TEXT NOT NULL,
              thinking_level TEXT NOT NULL DEFAULT 'none', sections_json TEXT NOT NULL,
              input_tokens INTEGER, output_tokens INTEGER, cost_usd REAL,
              created_at TEXT DEFAULT (datetime('now', 'localtime')));
            INSERT INTO sessions(id,date) VALUES(1,'2026-08-12');
            INSERT INTO runs(id,session_id,run_number,source_file)
              VALUES(1,1,1,'PROJ0001'),(2,1,2,'PROJ0002');
            INSERT INTO generated_reports(id,session_date,run_number,run_label,model,sections_json)
              VALUES(1,'2026-08-12',2,'Run 2','m','{}'),
                    (2,'2026-08-12',NULL,'All runs','m','{}');"""
        _p = _build('old_reports.db', _OLD_GR)
        _err = _migrate_only(_p)
        check(_err is None, 'pre-provenance generated_reports migrates', repr(_err))
        _cx = _sq.connect(_p)
        _cols = {r[1] for r in _cx.execute('PRAGMA table_info(generated_reports)')}
        check({'source_file', 'input_snapshot_json'} <= _cols, 'the two columns are added', str(sorted(_cols)))
        _rows = _cx.execute('SELECT id, source_file, input_snapshot_json FROM generated_reports ORDER BY id').fetchall()
        check(_rows[0][1] == 'PROJ0002', 'a single-run row is backfilled by position', str(_rows))
        check(_rows[1][1] is None, 'a whole-session row stays unpinned', str(_rows))
        check(all(r[2] is None for r in _rows), 'no snapshot is invented for old rows')
        _cx.close()
        _err = _migrate_only(_p)
        check(_err is None, 'and the migration is idempotent', repr(_err))
        _cx = _sq.connect(_p)
        check(_cx.execute('SELECT source_file FROM generated_reports WHERE id=1').fetchone()[0] == 'PROJ0002',
              'a second run leaves the backfill alone')
        _cx.close()

        # The clients send the identity alongside the number.
        _idx_src = open(os.path.join(REPO, 'templates', 'index.html')).read()
        check('body.source_file = _roptSrc' in _idx_src and 'proj.source_file);' in _idx_src,
              'index.html posts source_file with run_number')
        _rep_src = open(os.path.join(REPO, 'templates', 'reports.html')).read()
        check('body.source_file = src' in _rep_src, 'reports.html posts source_file with run_number')

        # ── 12. WP6 — serving and operations (F16, F17) ────────────────────────
        print('\n12. serving and operations (F16/F17)')
        import subprocess as _sp6

        # 11a. init_db() (all migrations) runs at import time now, not only
        # under `if __name__ == '__main__':` — gunicorn imports this module
        # and calls the WSGI `app` object directly, it never executes
        # __main__, so a fresh Pi used to come up serving requests against a
        # database with no tables at all. A real interpreter, a fresh
        # NOISE_DB_PATH nothing else has touched, and nothing but `import
        # noise_app` — the decisive version of this check, same shape as the
        # startup-security one in section 9a.
        _wp6_fresh_db = os.path.join(tmp, 'wp6-fresh.db')
        _wp6_env = dict(os.environ)
        _wp6_env['NOISE_DB_PATH'] = _wp6_fresh_db
        _wp6_check = (
            "import noise_app, sqlite3\n"
            "c = sqlite3.connect(%r)\n"
            "names = {r[0] for r in c.execute(\"SELECT name FROM sqlite_master WHERE type='table'\")}\n"
            "assert {'sessions', 'runs'} <= names, names\n"
        ) % _wp6_fresh_db
        _wp6_proc = _sp6.run([sys.executable, '-c', _wp6_check], cwd=REPO,
                             env=_wp6_env, capture_output=True, text=True)
        check(_wp6_proc.returncode == 0,
              'importing noise_app alone creates the schema (no __main__ needed)',
              (_wp6_proc.stderr or '').strip()[-300:])

        # 11b. get_db() puts the database in WAL mode — readers and writers
        # (request threads, the weather-fetch thread, the peer-push thread)
        # no longer block each other the way the default rollback journal did.
        sys.modules.pop('noise_app', None)
        wp6 = Side(os.path.join(tmp, 'wp6.db'))
        _mode = wp6.sql('PRAGMA journal_mode')[0][0]
        check(str(_mode).lower() == 'wal', 'get_db() puts the database in WAL mode', _mode)

        # 11c/d. MAX_CONTENT_LENGTH is set (env-overridable — a small value
        # here so the test doesn't need to build a real 64 MB body), and a
        # request over it gets a 413 with a readable body instead of
        # Werkzeug's bare "413 Request Entity Too Large" error page.
        os.environ['MAX_CONTENT_LENGTH_MB'] = '1'
        try:
            sys.modules.pop('noise_app', None)
            wp6app = importlib.import_module('noise_app')
            check(wp6app.app.config['MAX_CONTENT_LENGTH'] == 1024 * 1024,
                  'MAX_CONTENT_LENGTH honours MAX_CONTENT_LENGTH_MB',
                  wp6app.app.config['MAX_CONTENT_LENGTH'])
            wp6app.app.config['TESTING'] = True
            wp6cl = wp6app.app.test_client()
            _big = b'x' * (2 * 1024 * 1024)
            _resp413 = wp6cl.post('/upload', data={'file': (_io.BytesIO(_big), 'big.zip')},
                                  content_type='multipart/form-data')
            check(_resp413.status_code == 413,
                  'a request over MAX_CONTENT_LENGTH gets 413', _resp413.status_code)
            check(b'too large' in _resp413.get_data().lower(),
                  'and the 413 body is readable, not a bare Werkzeug error page',
                  _resp413.get_data()[:200])
            # The machine clients (import_sdcard.py, the peer Pi) get JSON,
            # not an HTML page, on the same failure.
            _resp413b = wp6cl.post('/import', data={'file': (_io.BytesIO(_big), 'big.zip')},
                                   content_type='multipart/form-data',
                                   headers={'X-Import-Key': 'whatever'})
            check(_resp413b.status_code == 413 and
                  'too large' in (_resp413b.get_json() or {}).get('message', '').lower(),
                  '/import gets the same 413 as JSON', _resp413b.get_data()[:200])

            # 11e. /health reports a version, falling back to 'dev' with no
            # VERSION file — there is none in this checkout; deploy_to_pis.py
            # writes one after copying files on a real Pi.
            _health = wp6cl.get('/health').get_json()
            check(_health.get('status') == 'ok' and _health.get('version') == 'dev',
                  "/health reports status and version ('dev' with no VERSION file)",
                  _health)
        finally:
            os.environ.pop('MAX_CONTENT_LENGTH_MB', None)

        # ── 13. WP5 — calculation conventions (F8, F10, F11, F12, F13) ────────
        print('\n13. session LAeq is duration-weighted (F8 / decision D2)')

        # A session whose runs are deliberately of very different lengths, with
        # round LAeq values so the expected figure can be computed by hand:
        #   run A: LAeq 70.0 dB over  100 s
        #   run B: LAeq 80.0 dB over  900 s
        # duration-weighted: 10*log10((100*10^7 + 900*10^8)/1000) = 79.5904... dB
        #        equal-weight: 10*log10((10^7 + 10^8)/2)          = 77.4036... dB
        # 2.2 dB apart, so the two conventions cannot be confused for each other.
        # n_samples is left at 900 for BOTH runs: if the weight came from the
        # record count rather than the meter's duration_s, the answer would be
        # the equal-weight one and this section would fail.
        WD = '2026-08-13'
        w = Side(os.path.join(tmp, 'weighted.db'))
        wsess = _copy.deepcopy(sessions[0])
        wsess['d'] = WD
        wsess['complete_date'] = False
        wsess['skipped_files'] = []
        wsess['projects'] = _copy.deepcopy(sessions[0]['projects'][:2])
        for _p, (_avg, _dur) in zip(wsess['projects'], ((70.0, 100), (80.0, 900))):
            _p['avg'], _p['duration_s'], _p['n'] = _avg, _dur, 900
        w.db.import_sessions([wsess], metadata=META)

        EXPECT_W = 79.59      # hand-computed above, half-up to 2 dp
        EXPECT_FLAT = 77.4    # what the equal-weight mean would have given
        wrow = w.sql('SELECT avg_laeq, run_count FROM sessions WHERE date=?', WD)[0]
        check(wrow['avg_laeq'] == EXPECT_W,
              'import_sessions stores the duration-weighted session LAeq',
              f'{wrow["avg_laeq"]} vs {EXPECT_W}')
        check(abs(wrow['avg_laeq'] - EXPECT_FLAT) > 2.0,
              'and it is not the equal-weight mean of the two run LAeqs',
              f'{wrow["avg_laeq"]} vs {EXPECT_FLAT}')

        # The same number must come back out of a recompute, or a backfill would
        # silently move every session it touched.
        check(w.db.recompute_session_aggregates([WD]) == [],
              'recompute agrees with import — nothing to change')
        w.exec('UPDATE sessions SET avg_laeq=1.0 WHERE date=?', WD)
        w.db.recompute_session_aggregates([WD])
        check(w.sql('SELECT avg_laeq FROM sessions WHERE date=?', WD)[0]['avg_laeq'] == EXPECT_W,
              'recompute_session_aggregates rebuilds the same weighted value',
              str(EXPECT_W))

        # And the report block, which is where the weighted figure already came
        # from, must now be the same number to its own (1 dp) rounding.
        _prep = w.reports._prepare_session_for_report(WD)
        check(_prep is not None, 'the weighted session prepares for a report')
        _ws, _wrows, _wall, _wtot = _prep
        _, _wstats = w.reports._build_session_data_block(_ws, _wrows, _wall, _wtot)
        check(_wstats['session_leq'] == _rhu(EXPECT_W, 1),
              'the report block agrees with the stored session LAeq',
              f'{_wstats["session_leq"]} vs {_rhu(EXPECT_W, 1)}')
        check(_wtot == 1000, 'total duration sums the real duration_s', str(_wtot))
        check([r.get('duration_s') for r in _wrows] == [100, 900],
              'run rows carry duration_s for the weighting',
              str([r.get('duration_s') for r in _wrows]))

        # duration_s is the preferred weight, n_samples the fallback. With the
        # durations removed, the same two runs weight equally (n is 900 for
        # both) and the session LAeq becomes the equal-weight figure.
        w.exec('UPDATE runs SET duration_s=NULL WHERE session_id='
               '(SELECT id FROM sessions WHERE date=?)', WD)
        w.db.recompute_session_aggregates([WD])
        _fallback = w.sql('SELECT avg_laeq FROM sessions WHERE date=?', WD)[0]['avg_laeq']
        check(abs(_fallback - EXPECT_FLAT) < 0.01,
              'with no duration_s the weight falls back to n_samples',
              f'{_fallback} vs {EXPECT_FLAT}')

        print('\n13b. BS 4142 has two periods, and boundary crossings are flagged (F10)')
        _tp = w.assess._time_period
        check(_tp('07:00:00', 'bs4142') == 'Day', 'BS 4142: 07:00 is Day')
        check(_tp('22:59:59', 'bs4142') == 'Day', 'BS 4142: 22:59 is still Day')
        check(_tp('23:00:00', 'bs4142') == 'Night', 'BS 4142: 23:00 is Night')
        check(_tp('06:59:59', 'bs4142') == 'Night', 'BS 4142: 06:59 is Night')
        # The defect: 19:00–23:00 used to be reported as a third period the
        # standard does not define.
        check(_tp('19:00:00', 'bs4142') == 'Day',
              'BS 4142: 19:00 is Day, not the old "Evening"', _tp('19:00:00', 'bs4142'))
        _all_bs = {_tp(f'{h:02d}:30:00', 'bs4142') for h in range(24)}
        check(_all_bs == {'Day', 'Night'},
              'BS 4142 yields only Day and Night, at every hour of the day', str(_all_bs))
        # assessments.html carries its own copy of this classification for the
        # run chips, so the page and the CSV would otherwise disagree.
        _ah = open(os.path.join(REPO, 'templates', 'assessments.html'), encoding='utf-8').read()
        _ajs = _ah.split('function timePeriod(')[1].split('\nfunction ')[0]
        check("'Evening'" not in _ajs,
              'the assessments.html copy of timePeriod() has no Evening either')
        check("['Day','Night']" in _ah,
              'and the BS 4142 period list on the page is Day/Night')
        check('Day/Evening/Night' not in _ah,
              'the standard picker no longer advertises three BS 4142 periods')
        # The Noise Act split is a different scheme and is unchanged.
        check(_tp('19:00:00', 'noise_act') == 'Pre 23:00', 'Noise Act: 19:00 is Pre 23:00')
        check(_tp('23:30:00', 'noise_act') == 'Post 23:00', 'Noise Act: 23:30 is Post 23:00')
        check(_tp('nonsense', 'bs4142') == 'unknown', 'an unparseable start time is unknown')

        _sb = w.assess._spans_boundary
        check(_sb('22:30:00', '23:30:00') is True, 'a run over 23:00 spans a boundary')
        check(_sb('06:30:00', '07:30:00') is True, 'a run over 07:00 spans a boundary')
        check(_sb('07:30:00', '08:30:00') is False, 'a run inside the day does not')
        check(_sb('22:00:00', '23:00:00') is False,
              'a run ending exactly on the boundary has not crossed it')
        check(_sb('23:30:00', '07:30:00') is True,
              'a run through midnight and past 07:00 spans a boundary')
        check(_sb('23:30:00', '06:00:00') is False,
              'a run through midnight that stops before 07:00 does not')
        # No arithmetic fallback: without the meter's end time this is unknown,
        # and unknown is reported as such rather than guessed from n_samples.
        check(_sb('22:30:00', None) is None, 'no end time means the answer is None')
        check(_sb('22:30:00', '') is None, 'an empty end time is also None')

        print('\n13c. assessment fallback percentiles read the stored 1 s series (F11)')
        # A run with no GLOB percentiles — the pre-backfill case the fallback
        # exists for. Everything else about run 9 is left alone.
        af = Side(os.path.join(tmp, 'fallback.db'))
        af.db.import_sessions(_copy.deepcopy(sessions), metadata=META)
        aid11 = af.assess.create_assessment('F11 fallback', standard='bs4142')
        lid11 = af.assess.add_assessment_location(aid11, 'L')
        af.assess.assign_runs(aid11, lid11, [(SESSION_DATE, 9)])
        _r9 = af.db.get_full_run_row(SESSION_DATE, 9)
        _series = json.loads(_r9['prof_lafspl_json'])
        _chart = json.loads(_r9['laeq_json'])
        af.exec('UPDATE runs SET la_l10=NULL, la_l50=NULL, la_l90=NULL WHERE id=?', _r9['id'])

        _fb = af.assess.prepare_assessment_report_data(aid11)['locations'][0]['runs'][0]
        _sv = sorted(_series)
        check(_fb['la10'] == af.db.percentile(_sv, 90),
              'fallback LA10 is the interpolated 90th percentile of prof_lafspl_json',
              f'{_fb["la10"]} vs {af.db.percentile(_sv, 90)}')
        check(_fb['la50'] == af.db.percentile(_sv, 50), 'fallback LA50 likewise')
        check(_fb['la90'] == af.db.percentile(_sv, 10),
              'fallback LA90 is the 10th percentile (the level exceeded 90% of the time)')
        check(_fb['la90'] < _fb['la50'] < _fb['la10'], 'fallback LA90 < LA50 < LA10',
              f'{_fb["la90"]} / {_fb["la50"]} / {_fb["la10"]}')
        # The old computation: nearest-rank over laeq_json, the downsampled chart
        # profile whose points are window maxima. It read high, and disagreed
        # with the figure the report gave for the same run.
        _old_sv = sorted(_chart, reverse=True)
        _old_la10 = round(_old_sv[max(0, int(len(_old_sv) * 0.1) - 1)], 1)
        check(_fb['la10'] != _old_la10,
              'and it is not the old chart-profile nearest-rank value',
              f'now {_fb["la10"]}, was {_old_la10}')
        check(_fb['la10'] < _old_la10,
              'the chart profile read high, as windows of maxima must',
              f'{_fb["la10"]} < {_old_la10}')
        # One implementation, so an assessment and a report cannot disagree.
        check(af.reports._percentile is af.db.percentile,
              'reports and assessments share one percentile implementation')
        check(af.reports._percentile(_sv, 90) == _fb['la10'],
              'the report and the assessment give the same LA10 for the run')
        # No stored series and no GLOB percentiles: report nothing, guess nothing.
        af.exec('UPDATE runs SET prof_lafspl_json=NULL WHERE id=?', _r9['id'])
        _none = af.assess.prepare_assessment_report_data(aid11)['locations'][0]['runs'][0]
        check(_none['la10'] is None and _none['la50'] is None and _none['la90'] is None,
              'with neither source the percentiles are None, not an estimate',
              str([_none['la10'], _none['la50'], _none['la90']]))

        print('\n13d. model pricing and per-model parameters (F12)')
        _P = w.reports._MODEL_PRICING
        check(_P['claude-haiku-4-5'] == (1.00, 5.00),
              'Haiku 4.5 is $1.00/$5.00 per MTok (was 0.80/4.00)', str(_P['claude-haiku-4-5']))
        check(_P['claude-sonnet-5'] == (3.00, 15.00),
              'Sonnet 5 is $3.00/$15.00 per MTok', str(_P['claude-sonnet-5']))
        check(_P['claude-opus-5'] == (5.00, 25.00),
              'Opus 5 is $5.00/$25.00 per MTok (was 15.00/75.00, an Opus 4.x figure)',
              str(_P['claude-opus-5']))
        check(_P['claude-haiku-4-5-20251001'] == _P['claude-haiku-4-5'],
              'the dated Haiku id is kept as an alias, so stored reports still price')
        for _mid in ('claude-haiku-4-5', 'claude-sonnet-5', 'claude-opus-5'):
            check(_mid in _P, f'canonical id {_mid} is priced')
        # The dropdowns must send ids the pricing table knows.
        for _tpl in ('index.html', 'reports.html'):
            _h = open(os.path.join(REPO, 'templates', _tpl), encoding='utf-8').read()
            _opts = [l.split('value="')[1].split('"')[0] for l in _h.split('\n')
                     if '<option value="claude-' in l]
            check(_opts and all(o in _P for o in _opts),
                  f'{_tpl}: every model option is a priced id', str(_opts))
        _rep_html = open(os.path.join(REPO, 'templates', 'report.html'), encoding='utf-8').read()
        check('API usage (claude-sonnet-5)' not in _rep_html,
              'report footer no longer names claude-sonnet-5 whatever model ran')
        check('usage_info.model' in _rep_html,
              'report footer names the model that actually generated it')

        # A stand-in for the anthropic SDK: records the kwargs it was called
        # with, so the per-model parameter rules can be asserted without a
        # network call or an API key.
        class _FakeBadRequest(Exception):
            pass

        _calls = []

        class _FakeAnthropic:
            def __init__(self, **kw):
                self.messages = self

            def create(self, **kwargs):
                _calls.append(kwargs)
                # Reproduce the real 400: Haiku 4.5 accepts neither parameter.
                if kwargs['model'].startswith('claude-haiku') and (
                        'thinking' in kwargs or 'output_config' in kwargs):
                    raise _FakeBadRequest(
                        'thinking.type: Input should be \'enabled\' or \'disabled\'')

                class _Block:
                    type = 'tool_use'
                    input = {k: '<p>x</p>' for k in SECTIONS}

                class _Usage:
                    input_tokens, output_tokens = 1_000_000, 1_000_000

                class _Msg:
                    content = [_Block()]
                    usage = _Usage()
                return _Msg()

        _fake_mod = types.ModuleType('anthropic')
        _fake_mod.Anthropic = _FakeAnthropic
        _fake_mod.BadRequestError = _FakeBadRequest
        _real_anthropic = sys.modules.get('anthropic')
        _real_key = os.environ.get('ANTHROPIC_API_KEY')
        sys.modules['anthropic'] = _fake_mod
        os.environ['ANTHROPIC_API_KEY'] = 'test-key-not-a-real-one'
        try:
            _calls.clear()
            _sections, _ti, _to, _cost = w.reports._call_claude('p', 'claude-sonnet-5', 'standard')
            check('thinking' in _calls[0] and 'output_config' in _calls[0],
                  'Sonnet 5 still gets adaptive thinking and an effort setting',
                  str(sorted(_calls[0])))
            # 1M in + 1M out at 3.00/15.00 = $18.0000 exactly.
            check(_cost == 18.0, 'cost is priced from the table', str(_cost))

            _calls.clear()
            _sections, _ti, _to, _cost = w.reports._call_claude('p', 'claude-haiku-4-5', 'standard')
            check('thinking' not in _calls[0] and 'output_config' not in _calls[0],
                  'Haiku 4.5 is sent neither thinking nor output_config (both are 400s)',
                  str(sorted(_calls[0])))
            check(_cost == 6.0, 'and is priced at 1.00/5.00', str(_cost))
            _calls.clear()
            w.reports._call_claude('p', 'claude-haiku-4-5', 'extended')
            check('thinking' not in _calls[0],
                  'the extended level is dropped for Haiku too, not just standard')

            check(w.reports._supports_thinking('claude-sonnet-5')
                  and not w.reports._supports_thinking('claude-haiku-4-5'),
                  '_supports_thinking classifies the two families')
        finally:
            if _real_anthropic is not None:
                sys.modules['anthropic'] = _real_anthropic
            else:
                sys.modules.pop('anthropic', None)
            if _real_key is None:
                os.environ.pop('ANTHROPIC_API_KEY', None)
            else:
                os.environ['ANTHROPIC_API_KEY'] = _real_key

        # The route maps a BadRequestError to 400 and anything else to 500.
        _errs = []

        def _boom(exc):
            def _f(prompt, model, thinking_level):
                raise exc
            return _f

        _real_call = rt.reports._call_claude
        _real_anthropic = sys.modules.get('anthropic')
        sys.modules['anthropic'] = _fake_mod
        try:
            _tid = rt.reports_db.save_report_template(
                'WP5 error-path template', '', 'x {{session_data}} y')
            _body = {'session_date': SESSION_DATE, 'template_id': _tid,
                     'model': 'claude-haiku-4-5'}
            rt.reports._call_claude = _boom(_FakeBadRequest('unsupported parameter'))
            _resp = _client.post('/api/generate-report', json=_body, headers=CSRF_HDR)
            check(_resp.status_code == 400,
                  'a rejected request returns 400, not 500', str(_resp.status_code))
            check('unsupported parameter' in _resp.get_json()['error'],
                  'and the API message reaches the caller')
            rt.reports._call_claude = _boom(RuntimeError('the roof is on fire'))
            _resp = _client.post('/api/generate-report', json=_body, headers=CSRF_HDR)
            check(_resp.status_code == 500,
                  'any other failure is still a 500', str(_resp.status_code))
        finally:
            rt.reports._call_claude = _real_call
            if _real_anthropic is not None:
                sys.modules['anthropic'] = _real_anthropic
            else:
                sys.modules.pop('anthropic', None)

        print('\n13e. one rounding convention for displayed dB (F13)')
        # 72.25 is exactly representable, so this is a genuine .x5 boundary and
        # not a float artefact: half-up gives 72.3, Python's banker's round 72.2.
        check(round(72.25, 1) == 72.2, 'Python round() is banker\'s at .x5 (the thing to avoid)')
        check(_rhu(72.25, 1) == 72.3, 'half-up rounds .x5 away from zero')
        check(w.db.percentile([72.0, 72.5], 50) == 72.3,
              'percentile() rounds half-up at a .x5 boundary (banker\'s would give 72.2)',
              str(w.db.percentile([72.0, 72.5], 50)))
        check(af.reports._percentile([72.0, 72.5], 50) == 72.3,
              'and so does the reports alias')
        check(w.db.percentile([], 50) is None, 'percentile of nothing is None')
        check(w.db.percentile([42.0], 90) == 42.0, 'percentile of one value is that value')

        # The same boundary through the assessment display path.
        _rb = Side(os.path.join(tmp, 'rounding.db'))
        _rb.db.import_sessions(_copy.deepcopy(sessions), metadata=META)
        _aid = _rb.assess.create_assessment('F13 rounding', standard='bs4142')
        _lid = _rb.assess.add_assessment_location(_aid, 'L')
        _rb.assess.assign_runs(_aid, _lid, [(SESSION_DATE, 9)])
        _rb.exec('UPDATE runs SET la_l10=72.25, avg_laeq=65.35 WHERE id='
                 '(SELECT r.id FROM runs r JOIN sessions s ON s.id=r.session_id '
                 ' WHERE s.date=? AND r.run_number=9)', SESSION_DATE)
        _rr = _rb.assess.prepare_assessment_report_data(_aid)['locations'][0]['runs'][0]
        check(_rr['la10'] == 72.3, 'assessment LA10 rounds half-up', str(_rr['la10']))
        check(_rr['avg_laeq'] == 65.4, 'assessment LAeq rounds half-up', str(_rr['avg_laeq']))
        check(_rb.reports._run_stats({'avg': 72.25, 'n': 1})['leq'] == 72.3,
              'the report run stats round half-up too',
              str(_rb.reports._run_stats({'avg': 72.25, 'n': 1})['leq']))
        # The decoder's convention, now shared: nothing should still be using
        # Python's round() on a value headed for a report or a stored dB. Two
        # uses of round() are not dB and stay — a compass index and a USD cost —
        # so they are named here rather than left to widen the pattern.
        _NOT_DB = ('deg / 45', 'cost_usd')
        for _mod in ('reports.py', 'assessments_db.py', 'noise_parser.py'):
            _src = open(os.path.join(REPO, _mod), encoding='utf-8').read()
            _bankers = [l.strip() for l in _src.split('\n')
                        if 'round(' in l and 'round_half_up' not in l
                        and '_round' not in l and not l.strip().startswith('#')
                        and not any(x in l for x in _NOT_DB)]
            check(not _bankers, f'{_mod}: no bare round() left on a dB value',
                  str(_bankers[:2]))

        print('\n13f. the spectrum reconstruction is kept, but not run on every import')
        # The handoff verifies the spectral decode by energy-summing the stored
        # 1/3-octave Lfeq table with IEC weighting and comparing to the meter's
        # own LAeq/LCeq. That check still passes — it is just no longer computed
        # 36 times per run on a value nothing read.
        _spec9 = json.loads(_r9['spec_lfeq'])
        _recon = parser.spectrum_broadband('spec_lfeq', _spec9)
        close(_recon['laeq'], 70.8, 0.06,
              'Lfeq spectrum reconstructs run 9 LAeq = 70.8 (handoff)')
        close(_recon['lceq'], 78.7, 0.06,
              'Lfeq spectrum reconstructs run 9 LCeq = 78.7 (handoff)')
        check(abs(parser._F4_HZ - 12194.217) < 1e-6,
              'the IEC 61672 f4 pole is 12194.217 Hz, not the rounded 12200',
              str(parser._F4_HZ))
        _pd, _pr = parser._parse_session_files(
            open(os.path.join(meas_root, DATE_FOLDER, 'PART0000', 'PROJ0009',
                              'GLOB0009.DAT'), 'rb').read(),
            open(os.path.join(meas_root, DATE_FOLDER, 'PART0000', 'PROJ0009',
                              'PROF0009.DAT'), 'rb').read())
        check(not [k for k in _pr if k.endswith('_from_spectrum')],
              'no *_from_spectrum value is computed on import any more',
              str([k for k in _pr if k.endswith('_from_spectrum')]))

        print('\n13g. the assessment CSV carries the boundary flag (F10)')
        _aid12 = rt.assess.create_assessment('F10 CSV', standard='bs4142')
        _lid12 = rt.assess.add_assessment_location(_aid12, 'L')
        rt.assess.assign_runs(_aid12, _lid12, [(SESSION_DATE, 9)])
        _resp = _client.get(f'/export/assessment/{_aid12}.csv')
        check(_resp.status_code == 200, 'assessment CSV downloads', _resp.status_code)
        _lines = _resp.get_data(as_text=True).split('\r\n')
        _hdr = _lines[0].split(',')
        check('spans_boundary' in _hdr, 'assessment CSV has a spans_boundary column', str(_hdr))
        _row = _lines[1].split(',')
        check(len(_row) == len(_hdr), 'CSV data row width still matches the header',
              f'{len(_row)} vs {len(_hdr)}')
        # Run 9 is 23:27:36–23:42:36, wholly inside the night period.
        check(_row[_hdr.index('time_period')] == 'Night', 'run 9 is a Night measurement',
              _row[_hdr.index('time_period')])
        check(_row[_hdr.index('spans_boundary')] == 'no', 'and it crosses no boundary',
              _row[_hdr.index('spans_boundary')])
        rt.assess.delete_assessment(_aid12)


        # ── 14. sessions keyed on (date, instrument_serial) (WP8) ────────────
        print('\n14. sessions are keyed on (date, instrument_serial)')
        # Several meters can measure on one calendar date. The serial comes
        # from outside the data (the NOR140 writes none into GLOB/PROF), rides
        # beside 'd' in every payload, and a payload naming none is filed
        # under the Pi's default serial — so everything pre-WP8 still imports.
        import io as _io14
        import zipfile as _zf14

        w = Side(os.path.join(tmp, 'wp8.db'))

        def _payload_copy(serial=None, bump=None, n_projects=2):
            sess14 = json.loads(json.dumps(dict(sessions[0])))
            sess14.pop('complete_date', None)   # partial payloads only add/refresh
            sess14['projects'] = sess14['projects'][:n_projects]
            if serial is None:
                sess14.pop('serial', None)   # a pre-WP8 payload has no such key
            else:
                sess14['serial'] = serial
            if bump:
                for pr in sess14['projects']:
                    pr['avg'] += bump
            return sess14

        # (a) two meters, same date, the same PROJ folder names — side by side
        sA = _payload_copy('METER-A')
        sB = _payload_copy('METER-B', bump=3.0)
        w.db.import_sessions([sA])
        w.db.import_sessions([sB])
        rows14 = w.sql('SELECT id, instrument_serial, run_count, avg_laeq '
                       'FROM sessions WHERE date=? ORDER BY instrument_serial',
                       SESSION_DATE)
        check([r['instrument_serial'] for r in rows14] == ['METER-A', 'METER-B'],
              'two meters on one date are two sessions',
              str([dict(r) for r in rows14]))
        check(all(r['run_count'] == 2 for r in rows14),
              'each meter keeps its own runs', str([r['run_count'] for r in rows14]))
        check(rows14[0]['avg_laeq'] != rows14[1]['avg_laeq'],
              'aggregates are computed per meter, not pooled',
              f"{rows14[0]['avg_laeq']} vs {rows14[1]['avg_laeq']}")
        for r in rows14:
            nums = [x['run_number'] for x in w.sql(
                'SELECT run_number FROM runs WHERE session_id=? ORDER BY run_number', r['id'])]
            check(nums == [1, 2], f"{r['instrument_serial']}: run numbering starts at 1",
                  str(nums))
        w.db.import_sessions([sA])   # re-import upserts, never twins
        check(len(w.sql('SELECT id FROM sessions WHERE date=?', SESSION_DATE)) == 2,
              'a re-import lands on the same (date, serial) session')
        srcA = w.sql('SELECT r.source_file FROM runs r JOIN sessions s ON s.id=r.session_id '
                     "WHERE s.date=? AND s.instrument_serial='METER-A' AND r.run_number=1",
                     SESSION_DATE)[0]['source_file']
        srcB = w.sql('SELECT r.source_file FROM runs r JOIN sessions s ON s.id=r.session_id '
                     "WHERE s.date=? AND s.instrument_serial='METER-B' AND r.run_number=1",
                     SESSION_DATE)[0]['source_file']
        check(srcA == srcB, 'the same PROJ folder can exist under both serials', f'{srcA} / {srcB}')
        rrA = w.db.get_full_run_row(SESSION_DATE, 1, serial='METER-A')
        rrB = w.db.get_full_run_row(SESSION_DATE, 1, serial='METER-B')
        check(abs(rrB['avg_laeq'] - rrA['avg_laeq'] - 3.0) < 1e-6,
              'get_full_run_row(serial=…) reads the right meter',
              f"{rrA['avg_laeq']} vs {rrB['avg_laeq']}")

        # (b) an old-format payload (no serial) lands on the default serial
        w.db.set_setting('instrument_serial', '6899108')
        w.db.import_sessions([_payload_copy(serial=None)])
        check(bool(w.sql('SELECT 1 FROM sessions WHERE date=? AND instrument_serial=?',
                         SESSION_DATE, '6899108')),
              'a payload without serial is filed under the instrument_serial setting')
        check(len(w.sql('SELECT id FROM sessions WHERE date=?', SESSION_DATE)) == 3,
              'and is a third session beside the two named meters')

        # (c) the serial survives a peer round-trip
        payload14 = w.db.get_sessions_since('1970-01-01T00:00:00')
        check(sorted(p14['serial'] for p14 in payload14) == ['6899108', 'METER-A', 'METER-B'],
              'the sync payload carries serial beside d',
              str(sorted(p14.get('serial') for p14 in payload14)))
        w2 = Side(os.path.join(tmp, 'wp8-peer.db'))
        w2.db.import_sessions(payload14)
        check(sorted(r['instrument_serial'] for r in
                     w2.sql('SELECT instrument_serial FROM sessions WHERE date=?', SESSION_DATE))
              == ['6899108', 'METER-A', 'METER-B'],
              'the peer stores each session under the sender\'s serial')

        # (e, first half) an assessment link and a generated report name the meter
        aid14 = w.assess.create_assessment('WP8', standard='bs4142')
        lid14 = w.assess.add_assessment_location(aid14, 'L')
        w.assess.assign_runs(aid14, lid14, [(SESSION_DATE, 1, srcB, 'METER-B')])
        det14 = w.assess.get_assessment_detail(aid14)
        _linked = det14['locations'][0]['runs']
        check(len(_linked) == 1 and abs(_linked[0]['avg_laeq'] - rrB['avg_laeq']) < 1e-6,
              'the link resolves to METER-B\'s run, not METER-A\'s same-named one',
              str([(_r['instrument_serial'], _r['avg_laeq']) for _r in _linked]))
        pool14 = w.assess.get_all_runs_for_assessment(aid14)
        marked = [(r14['instrument_serial'], r14['run_number'])
                  for r14 in pool14 if r14['ar_id'] is not None]
        check(marked == [('METER-B', 1)], 'the pool marks only METER-B run 1', str(marked))
        rid14 = w.reports_db.save_generated_report(
            SESSION_DATE, None, 't', 'claude-sonnet-5', 'none', '{}', 1, 1, 0.0,
            run_number=1, source_file=srcB, instrument_serial='METER-B')
        check(w.reports_db.get_generated_report(rid14)['instrument_serial'] == 'METER-B',
              'a generated report records the instrument')

        # (d) tombstones carry the serial; one without applies to the default
        w.db.delete_session(SESSION_DATE, 'METER-A')
        tombs14 = w.db.get_session_tombstones()
        check([(t['date'], t['serial']) for t in tombs14] == [(SESSION_DATE, 'METER-A')],
              'delete_session(date, serial) tombstones that meter only', str(tombs14))
        check(len(w.sql('SELECT id FROM sessions WHERE date=?', SESSION_DATE)) == 2,
              'the other two sessions on the date survive')

        w2.exec("UPDATE sessions SET imported_at='2020-01-01 00:00:00'")
        w2.sync.apply_full_sync(w.sync.get_full_sync_payload())
        left14 = sorted(r['instrument_serial'] for r in
                        w2.sql('SELECT instrument_serial FROM sessions WHERE date=?', SESSION_DATE))
        check(left14 == ['6899108', 'METER-B'],
              'the replicated tombstone deletes only METER-A on the peer', str(left14))
        # an old peer's tombstone names no serial → the default serial only
        w2.db.set_setting('instrument_serial', '6899108')
        w2.sync.apply_full_sync({'deleted_sessions': [
            {'date': SESSION_DATE, 'deleted_at': '2031-01-01 00:00:00'}]})
        left14 = sorted(r['instrument_serial'] for r in
                        w2.sql('SELECT instrument_serial FROM sessions WHERE date=?', SESSION_DATE))
        check(left14 == ['METER-B'],
              'a serial-less tombstone applies to the default serial only', str(left14))

        # (e, second half) the link survived METER-A\'s deletion, on the right meter
        det14 = w.assess.get_assessment_detail(aid14)
        check(len(det14['locations'][0]['runs']) == 1,
              'the assessment link survives the other meter\'s deletion')
        audit14 = w.db.audit_assessment_run_keys()
        check(not any(audit14.values()), 'and the key audit is clean', str(audit14))

        # session delete event without serial → the receiver\'s default
        w2.db.import_sessions([_payload_copy(serial='6899108')])
        w2.sync.apply_sync_event('session', 'delete', {'date': SESSION_DATE})
        left14 = sorted(r['instrument_serial'] for r in
                        w2.sql('SELECT instrument_serial FROM sessions WHERE date=?', SESSION_DATE))
        check(left14 == ['METER-B'],
              "a delete event without serial deletes the receiver's default serial",
              str(left14))

        # run_tag events: applied by source_file within the named serial
        w.sync.apply_sync_event('run_tag', 'upsert', {
            'session_date': SESSION_DATE, 'serial': 'METER-B',
            'run_number': 999, 'source_file': srcB, 'location_tag': 'Q'})
        check(w.db.get_full_run_row(SESSION_DATE, 1, serial='METER-B')['location_tag'] == 'Q',
              'a run_tag event lands by (serial, source_file), ignoring a stale number')

        # ── 14b. migration of a pre-WP8 database ──────────────────────────────
        print('\n14b. migration of a pre-WP8 database')
        _p14 = os.path.join(tmp, 'prewp8.db')
        _cx = _sq.connect(_p14)
        _cx.executescript("""
            CREATE TABLE sessions(id INTEGER PRIMARY KEY, date TEXT UNIQUE NOT NULL,
              run_count INTEGER, avg_laeq REAL, max_laeq REAL, recorder_name TEXT,
              location_label TEXT, postcode TEXT, lat REAL, lng REAL,
              imported_at TEXT DEFAULT (datetime('now')), notes TEXT);
            CREATE INDEX idx_sessions_date ON sessions(date);
            CREATE TABLE runs(id INTEGER PRIMARY KEY,
              session_id INTEGER REFERENCES sessions(id) ON DELETE CASCADE,
              run_number INTEGER, start_time TEXT, n_samples INTEGER,
              step INTEGER DEFAULT 1, avg_laeq REAL, min_laeq REAL, max_laeq REAL,
              max_lcpeak REAL, laeq_json TEXT, lcpeak_json TEXT, source_file TEXT,
              UNIQUE(session_id, run_number));
            CREATE TABLE deleted_sessions(date TEXT PRIMARY KEY,
              deleted_at TEXT DEFAULT (datetime('now')));
            CREATE TABLE assessments(id INTEGER PRIMARY KEY, name TEXT);
            CREATE TABLE assessment_locations(id INTEGER PRIMARY KEY, assessment_id INTEGER);
            CREATE TABLE assessment_runs(id INTEGER PRIMARY KEY,
              assessment_id INTEGER NOT NULL REFERENCES assessments(id) ON DELETE CASCADE,
              location_id INTEGER, session_date TEXT NOT NULL, run_number INTEGER,
              source_file TEXT, conditions TEXT, notes TEXT);
            CREATE UNIQUE INDEX idx_assessment_runs_stable
              ON assessment_runs(assessment_id, session_date, source_file)
              WHERE source_file IS NOT NULL;
            CREATE TABLE generated_reports(id INTEGER PRIMARY KEY, session_date TEXT NOT NULL,
              run_number INTEGER, run_label TEXT, template_id INTEGER, template_name TEXT,
              model TEXT NOT NULL, thinking_level TEXT NOT NULL DEFAULT 'none',
              sections_json TEXT NOT NULL, input_tokens INTEGER, output_tokens INTEGER,
              cost_usd REAL, created_at TEXT, source_file TEXT, input_snapshot_json TEXT);
            CREATE TABLE app_settings(key TEXT PRIMARY KEY, value TEXT);
            INSERT INTO app_settings VALUES('instrument_serial','6899108');
            INSERT INTO sessions(id,date,run_count,avg_laeq,max_laeq,notes)
              VALUES(7,'2026-08-12',2,55.5,71.0,'kept'),(9,'2026-08-13',1,50.0,60.0,NULL);
            INSERT INTO runs(id,session_id,run_number,start_time,avg_laeq,source_file)
              VALUES(1,7,1,'12:00:00',55.0,'PROJ0001'),(2,7,2,'13:00:00',56.0,'PROJ0002'),
                    (3,9,1,'09:00:00',50.0,'PROJ0001');
            INSERT INTO deleted_sessions VALUES('2026-07-01','2026-07-02 00:00:00');
            INSERT INTO assessments VALUES(1,'A');
            INSERT INTO assessment_runs(id,assessment_id,session_date,run_number,source_file)
              VALUES(4,1,'2026-08-12',2,'PROJ0002');
            INSERT INTO generated_reports(id,session_date,run_number,model,sections_json,source_file)
              VALUES(2,'2026-08-12',1,'claude-sonnet-5','{}','PROJ0001');
        """)
        _cx.commit(); _cx.close()

        _err = _migrate_only(_p14)
        check(_err is None, 'a pre-WP8 database migrates', repr(_err))
        _cx = _sq.connect(_p14); _cx.row_factory = _sq.Row
        _sess = [dict(r) for r in _cx.execute(
            'SELECT id, date, instrument_serial, run_count, avg_laeq, notes '
            'FROM sessions ORDER BY id').fetchall()]
        check([(r['id'], r['date'], r['instrument_serial']) for r in _sess]
              == [(7, '2026-08-12', '6899108'), (9, '2026-08-13', '6899108')],
              'session ids kept; every row backfilled with the default serial', str(_sess))
        check(_sess[0]['run_count'] == 2 and _sess[0]['avg_laeq'] == 55.5
              and _sess[0]['notes'] == 'kept',
              'session values survive the table rebuild', str(_sess[0]))
        _runs = _cx.execute(
            'SELECT COUNT(*) FROM runs r JOIN sessions s ON s.id=r.session_id').fetchone()[0]
        check(_runs == 3, 'every run still joins its session by id', str(_runs))
        check('UNIQUE(date, instrument_serial)' in _cx.execute(
            "SELECT sql FROM sqlite_master WHERE name='sessions'").fetchone()[0],
              'UNIQUE(date) became UNIQUE(date, instrument_serial)')
        _ds = [tuple(r) for r in _cx.execute(
            'SELECT date, instrument_serial, deleted_at FROM deleted_sessions').fetchall()]
        check(_ds == [('2026-07-01', '6899108', '2026-07-02 00:00:00')],
              'tombstones re-keyed to (date, serial) with times preserved', str(_ds))
        _ar = dict(_cx.execute('SELECT * FROM assessment_runs').fetchone())
        check(_ar['id'] == 4 and _ar['instrument_serial'] == '6899108'
              and _ar['source_file'] == 'PROJ0002',
              'assessment link id kept and backfilled with the serial', str(_ar))
        _gr = dict(_cx.execute('SELECT * FROM generated_reports WHERE id=2').fetchone())
        check(_gr['instrument_serial'] == '6899108',
              'generated report backfilled with the serial', str(_gr))
        _idx = _cx.execute("SELECT sql FROM sqlite_master WHERE name='idx_assessment_runs_stable'"
                           ).fetchone()[0]
        check('instrument_serial' in _idx, 'the stable unique index includes the serial', _idx)
        _before = {t: _cx.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0]
                   for t in ('sessions', 'runs', 'assessment_runs',
                             'generated_reports', 'deleted_sessions')}
        _cx.close()
        _err = _migrate_only(_p14)   # idempotency: a second migration is a no-op
        check(_err is None, 'the migration is idempotent', repr(_err))
        _cx = _sq.connect(_p14)
        _after = {t: _cx.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0]
                  for t in ('sessions', 'runs', 'assessment_runs',
                            'generated_reports', 'deleted_sessions')}
        check(_before == _after, 'and changes no row counts', f'{_before} vs {_after}')
        check(_cx.execute('SELECT id FROM sessions ORDER BY id').fetchall() == [(7,), (9,)],
              'and keeps the session ids a second time')
        _cx.close()

        # ── 14c. the routes: upload form serial, CSVs, deep-link statics ─────
        print('\n14c. routes and templates carry the serial')
        # The upload form field files the card under the named meter (route.db,
        # via the 4d client — its existing runs sit under the '' serial, so the
        # same start times under WP8-TEST must not be deduplicated away).
        _zbuf = _io14.BytesIO()
        with _zf14.ZipFile(_zbuf, 'w') as _z:
            for _fn in ('GLOB0004.DAT', 'PROF0004.DAT'):
                _fp = os.path.join(meas_root, DATE_FOLDER, 'PART0000', 'PROJ0004', _fn)
                _z.write(_fp, f'{DATE_FOLDER}/PART0000/PROJ0004/{_fn}')
        _resp = _client.post('/upload', headers=CSRF_HDR, data={
            'serial': 'WP8-TEST',
            'file': (_io14.BytesIO(_zbuf.getvalue()), 'card.zip'),
        }, content_type='multipart/form-data')
        check(_resp.status_code in (200, 302), 'upload with a serial field succeeds',
              f'{_resp.status_code}: {_resp.get_data(as_text=True)[:150]}')
        _n = rt.sql('SELECT COUNT(*) FROM sessions WHERE date=? AND instrument_serial=?',
                    SESSION_DATE, 'WP8-TEST')[0][0]
        check(_n == 1, 'the card landed under the form serial, beside the existing session')
        check(rt.db.get_setting('last_upload_serial') == 'WP8-TEST',
              'the serial is remembered as the next upload\'s default')
        _r = _client.get('/api/existing-dates?serial=WP8-TEST')
        check(SESSION_DATE in _r.get_json(), 'existing-dates?serial= sees that meter')
        _r = _client.get('/api/existing-dates?serial=SOME-OTHER')
        check(SESSION_DATE not in _r.get_json(), 'and not an unknown one')

        # sessions CSV carries a serial column
        _r = _client.get('/export/sessions.csv')
        _hdr14 = _r.get_data(as_text=True).split('\r\n')[0].split(',')
        check(_hdr14[:3] == ['date', 'serial', 'location'],
              'sessions CSV has the serial column', str(_hdr14[:3]))
        # assessment CSV too (header checked; 13g covered the values)
        _aid14 = rt.assess.create_assessment('WP8 hdr', standard='bs4142')
        _r = _client.get(f'/export/assessment/{_aid14}.csv')
        _hdr14 = _r.get_data(as_text=True).split('\r\n')[0].split(',')
        check('serial' in _hdr14, 'assessment CSV has a serial column', str(_hdr14))
        rt.assess.delete_assessment(_aid14)

        # import_sdcard: --serial flag exists and stamps the payload
        import import_sdcard as _isd14
        _sd14 = os.path.join(tmp, 'sd14', DATE_FOLDER, 'PART0000', 'PROJ0004')
        os.makedirs(_sd14)
        for _fn in ('GLOB0004.DAT', 'PROF0004.DAT'):
            shutil.copy(os.path.join(meas_root, DATE_FOLDER, 'PART0000', 'PROJ0004', _fn), _sd14)
        _sessions14 = _isd14.parse_all(sd_root=os.path.join(tmp, 'sd14'), serial='CLI-1')
        check(len(_sessions14) == 1 and _sessions14[0]['serial'] == 'CLI-1',
              'import_sdcard stamps the --serial into the payload',
              str([(x['d'], x.get('serial')) for x in _sessions14]))
        _isd_src = open(os.path.join(REPO, 'import_sdcard.py'), encoding='utf-8').read()
        check("add_argument('--serial'" in _isd_src and 'INSTRUMENT_SERIAL' in _isd_src,
              '--serial is a flag, defaulting from INSTRUMENT_SERIAL')

        # UI statics: the deep link reads &serial=, manage writes it, the
        # upload form names the field, and the report body sends it.
        _ih = open(os.path.join(REPO, 'templates', 'index.html'), encoding='utf-8').read()
        check("_urlParams.get('serial')" in _ih, 'index.html reads &serial= from the URL')
        check('selectedSerial' in _ih and 'MULTI_SERIAL' in _ih,
              'index.html tracks the selected serial and hides it while one meter exists')
        check("serial:         _roptSerial || ''" in _ih,
              'the report modal posts the serial')
        _mh = open(os.path.join(REPO, 'templates', 'manage.html'), encoding='utf-8').read()
        check('&serial={{ s.instrument_serial | urlencode }}' in _mh,
              'manage.html writes &serial= into View deep links')
        _uh = open(os.path.join(REPO, 'templates', 'upload.html'), encoding='utf-8').read()
        check('name="serial"' in _uh and 'upload_serial' in _uh,
              'upload.html has the serial field with the remembered default')
        _rh = open(os.path.join(REPO, 'templates', 'reports.html'), encoding='utf-8').read()
        check("serial: sess.serial || ''" in _rh,
              'reports.html sends the serial with generate-report')

        # ── 15. session-keyed weather; reports replicate (WP9) ───────────────
        print('\n15. weather is keyed to the session and replicates; reports replicate')
        # Weather belongs to the session — its time and location — so it is
        # keyed (date, instrument_serial) and replicates: the Pis are
        # dual-redundant, and so are generated reports (append-only evidence,
        # keyed across the pair by uid, never by local id). Report templates
        # stay per-Pi until F6.
        xa = Side(os.path.join(tmp, 'wp9-a.db'))
        _sA = _payload_copy('WX-A'); _sA['lat'], _sA['lng'] = 53.7997, -1.5492
        _sB = _payload_copy('WX-B'); _sB['lat'], _sB['lng'] = 51.5072, -0.1276
        xa.db.import_sessions([_sA, _sB])
        xa.db.save_weather(SESSION_DATE, {
            'wind_speed': 7.2, 'wind_dir': 210.0, 'temp_min': 11.0,
            'temp_max': 19.4, 'precip': 0.2,
            'hourly_json': '{"temperature_2m":[11.0]}'}, serial='WX-A')
        xa.db.save_weather(SESSION_DATE, {
            'wind_speed': 3.1, 'wind_dir': 90.0, 'temp_min': 14.0,
            'temp_max': 22.0, 'precip': 0.0,
            'hourly_json': '{"temperature_2m":[14.0]}'}, serial='WX-B')
        _wxrows = xa.sql('SELECT instrument_serial, wind_speed FROM weather '
                         'WHERE date=? ORDER BY instrument_serial', SESSION_DATE)
        check([tuple(r) for r in _wxrows] == [('WX-A', 7.2), ('WX-B', 3.1)],
              'two same-date sessions hold two distinct weather rows',
              str([tuple(r) for r in _wxrows]))
        check(xa.db.get_weather(SESSION_DATE, 'WX-A')['temp_max'] == 19.4
              and xa.db.get_weather(SESSION_DATE, 'WX-B')['temp_max'] == 22.0,
              'get_weather(date, serial) reads the right site')
        _all15 = xa.db.get_all_sessions_json()['sessions']
        _wxA = next(s for s in _all15 if s['d'] == SESSION_DATE and s['serial'] == 'WX-A')['wx']
        _wxB = next(s for s in _all15 if s['d'] == SESSION_DATE and s['serial'] == 'WX-B')['wx']
        check(_wxA['ws'] == 7.2 and _wxB['ws'] == 3.1,
              'each session card shows its own weather chips',
              f"{_wxA['ws']} / {_wxB['ws']}")
        check('hj' not in _wxA, 'the browser payload stays summary-only (no hourly blob)')

        # hourly_json replicates: sync payload carries it as wx.hj, the import
        # stores it, and an old-format payload (hj missing or None) cannot
        # erase a stored hourly series.
        _pl15 = xa.db.get_sessions_since('1970-01-01T00:00:00')
        _plA = next(p for p in _pl15 if p['serial'] == 'WX-A')
        check(_plA['wx'].get('hj') == '{"temperature_2m":[11.0]}',
              "the sync payload's wx dict carries hourly_json as 'hj'",
              str(_plA['wx']))
        xb = Side(os.path.join(tmp, 'wp9-b.db'))
        xb.db.import_sessions(json.loads(json.dumps(_pl15)))   # as the wire delivers it
        check(xb.db.get_weather(SESSION_DATE, 'WX-A')['hourly_json']
              == '{"temperature_2m":[11.0]}'
              and xb.db.get_weather(SESSION_DATE, 'WX-B')['hourly_json']
              == '{"temperature_2m":[14.0]}',
              'hourly_json survives the peer round-trip, per serial')
        _old15 = json.loads(json.dumps(_pl15))
        for _s15 in _old15:
            if _s15['serial'] == 'WX-A':
                _s15['wx'].pop('hj', None)      # pre-WP9 sender: no such key
            else:
                _s15['wx']['hj'] = None         # explicit None on the wire
            _s15['wx']['pr'] = 9.9              # but its summary is fresher
        xb.db.import_sessions(_old15)
        _after15 = {r['instrument_serial']: dict(r) for r in xb.sql(
            'SELECT * FROM weather WHERE date=?', SESSION_DATE)}
        check(_after15['WX-A']['hourly_json'] == '{"temperature_2m":[11.0]}'
              and _after15['WX-B']['hourly_json'] == '{"temperature_2m":[14.0]}',
              'an old-format payload does not erase the stored hourly series')
        check(_after15['WX-A']['precip'] == 9.9,
              'while its summary values still refresh the row')

        # A report saved on Side A reaches Side B through the mutation event,
        # applied by uid — the local ids differ by construction (the decoy
        # occupies xb's id 1, which is xa's id for the real report).
        _sections15 = json.dumps({'executive_summary': '<p>quiet</p>'})
        _snap15 = json.dumps({'stats': {'laeq': 55.5}, 'run_rows': []})
        rid15 = xa.reports_db.save_generated_report(
            SESSION_DATE, None, 'WP9 template', 'claude-sonnet-5', 'none',
            _sections15, 10, 20, 0.01, run_number=1, run_label='Run 1',
            source_file='PROJ0001', input_snapshot_json=_snap15,
            instrument_serial='WX-A')
        rowA15 = xa.reports_db.get_generated_report(rid15)
        check(bool(rowA15.get('uid')), 'a saved report carries a uid', str(rowA15.get('uid')))
        xb.reports_db.save_generated_report(
            SESSION_DATE, None, 'decoy', 'claude-sonnet-5', 'none', '{}', 1, 1, 0.0)
        _wire15 = json.loads(json.dumps(rowA15, default=str))  # as peer_client sends it
        xb.sync.apply_sync_event('generated_report', 'upsert', _wire15)
        _got15 = [dict(r) for r in xb.sql(
            'SELECT * FROM generated_reports WHERE uid=?', rowA15['uid'])]
        check(len(_got15) == 1, 'the event lands the report on the peer, keyed by uid')
        check(_got15[0]['sections_json'] == _sections15
              and _got15[0]['input_snapshot_json'] == _snap15,
              'with its sections and input snapshot intact')
        check(_got15[0]['id'] != rowA15['id'],
              'under a local id of its own — never the sender\'s',
              f"{_got15[0]['id']} vs {rowA15['id']}")
        xb.sync.apply_sync_event('generated_report', 'upsert', _wire15)
        check(xb.sql('SELECT COUNT(*) FROM generated_reports')[0][0] == 2,
              'a replayed event upserts rather than twins')

        # …and through the full-sync payload, weather rows included, so an
        # offline Pi catches up on startup. Templates stay out of it (F6).
        _fp15 = xa.sync.get_full_sync_payload()
        check('generated_reports' in _fp15 and 'weather' in _fp15,
              'the full-sync payload carries reports and weather')
        # WP10/F6: templates replicate now too, uid-keyed with LWW — this
        # asserted their absence while they were still per-Pi.
        check('report_templates' in _fp15,
              'and report templates (replicated by uid since WP10/F6)')
        zc = Side(os.path.join(tmp, 'wp9-full.db'))
        zc.sync.apply_full_sync(json.loads(json.dumps(_fp15, default=str)))
        _zrep = [dict(r) for r in zc.sql(
            'SELECT * FROM generated_reports WHERE uid=?', rowA15['uid'])]
        check(len(_zrep) == 1 and _zrep[0]['input_snapshot_json'] == _snap15,
              'full sync lands the report on a fresh Pi, snapshot intact')
        _zwx = zc.sql('SELECT instrument_serial, hourly_json FROM weather '
                      'WHERE date=? ORDER BY instrument_serial', SESSION_DATE)
        check([tuple(r) for r in _zwx] == [
                  ('WX-A', '{"temperature_2m":[11.0]}'),
                  ('WX-B', '{"temperature_2m":[14.0]}')],
              'full sync lands both weather rows, hourly included', str([tuple(r) for r in _zwx]))
        zc.sync.apply_full_sync(json.loads(json.dumps(_fp15, default=str)))
        check(zc.sql('SELECT COUNT(*) FROM generated_reports')[0][0] == 1
              and zc.sql('SELECT COUNT(*) FROM weather')[0][0] == 2,
              'a replayed full sync changes nothing')

        # deletes replicate by uid; a uid the receiver never had is a no-op
        xb.sync.apply_sync_event('generated_report', 'delete', {'uid': rowA15['uid']})
        check(xb.sql('SELECT COUNT(*) FROM generated_reports WHERE uid=?',
                     rowA15['uid'])[0][0] == 0,
              'a delete event removes the row by uid')
        check(xb.sql("SELECT template_name FROM generated_reports")[0][0] == 'decoy',
              'and leaves the peer\'s own local rows alone')
        xb.sync.apply_sync_event('generated_report', 'delete', {'uid': 'never-seen'})
        check(xb.sql('SELECT COUNT(*) FROM generated_reports')[0][0] == 1,
              'an unknown uid deletes nothing')

        # the routes are wired to the events (the client half of section 11
        # already generates over HTTP; here the mirroring is what matters)
        _rp15 = open(os.path.join(REPO, 'reports.py'), encoding='utf-8').read()
        check("sync_event_to_peer('generated_report', 'upsert', get_generated_report(rid))" in _rp15,
              'the generate route mirrors the saved row to the peer')
        check("sync_event_to_peer('generated_report', 'delete', info)" in _rp15,
              'the delete route mirrors the uid (with its tombstone time since WP10)')

        # ── 15b. migration of a pre-WP9 database ─────────────────────────────
        print('\n15b. migration of a pre-WP9 database')
        _p15b = os.path.join(tmp, 'prewp9.db')
        _cx = _sq.connect(_p15b)
        _cx.executescript("""
            CREATE TABLE sessions(id INTEGER PRIMARY KEY, date TEXT UNIQUE NOT NULL,
              run_count INTEGER, avg_laeq REAL, max_laeq REAL, recorder_name TEXT,
              location_label TEXT, postcode TEXT, lat REAL, lng REAL,
              imported_at TEXT DEFAULT (datetime('now')), notes TEXT);
            CREATE TABLE runs(id INTEGER PRIMARY KEY,
              session_id INTEGER REFERENCES sessions(id) ON DELETE CASCADE,
              run_number INTEGER, start_time TEXT, n_samples INTEGER,
              step INTEGER DEFAULT 1, avg_laeq REAL, min_laeq REAL, max_laeq REAL,
              max_lcpeak REAL, laeq_json TEXT, lcpeak_json TEXT, source_file TEXT,
              UNIQUE(session_id, run_number));
            CREATE TABLE deleted_sessions(date TEXT PRIMARY KEY,
              deleted_at TEXT DEFAULT (datetime('now')));
            CREATE TABLE weather(date TEXT PRIMARY KEY, wind_speed REAL,
              wind_dir REAL, temp_min REAL, temp_max REAL, precip REAL,
              hourly_json TEXT);
            CREATE TABLE generated_reports(id INTEGER PRIMARY KEY, session_date TEXT NOT NULL,
              run_number INTEGER, run_label TEXT, template_id INTEGER, template_name TEXT,
              model TEXT NOT NULL, thinking_level TEXT NOT NULL DEFAULT 'none',
              sections_json TEXT NOT NULL, input_tokens INTEGER, output_tokens INTEGER,
              cost_usd REAL, created_at TEXT, source_file TEXT, input_snapshot_json TEXT);
            CREATE TABLE app_settings(key TEXT PRIMARY KEY, value TEXT);
            INSERT INTO app_settings VALUES('instrument_serial','1402755');
            INSERT INTO sessions(id,date,run_count) VALUES(3,'2026-08-12',1);
            INSERT INTO runs(id,session_id,run_number,start_time,avg_laeq,source_file)
              VALUES(1,3,1,'12:00:00',55.0,'PROJ0001');
            INSERT INTO weather VALUES('2026-08-12',7.5,180.0,10.0,20.0,0.0,'{"h":1}'),
                                      ('2026-08-13',NULL,NULL,NULL,NULL,NULL,NULL);
            INSERT INTO generated_reports(id,session_date,run_number,model,sections_json)
              VALUES(1,'2026-08-12',1,'claude-sonnet-5','{}'),
                    (2,'2026-08-12',NULL,'claude-opus-4-6','{}');
        """)
        _cx.commit(); _cx.close()
        _err = _migrate_only(_p15b)
        check(_err is None, 'a pre-WP9 database migrates', repr(_err))
        _cx = _sq.connect(_p15b); _cx.row_factory = _sq.Row
        _wxm = [dict(r) for r in _cx.execute(
            'SELECT * FROM weather ORDER BY date').fetchall()]
        check([(r['date'], r['instrument_serial']) for r in _wxm]
              == [('2026-08-12', '1402755'), ('2026-08-13', '1402755')],
              'weather rows re-keyed under the default serial', str(_wxm))
        check(_wxm[0]['wind_speed'] == 7.5 and _wxm[0]['hourly_json'] == '{"h":1}',
              'weather values and the hourly blob survive the rebuild')
        check('PRIMARY KEY (date, instrument_serial)' in _cx.execute(
            "SELECT sql FROM sqlite_master WHERE name='weather'").fetchone()[0],
              'the weather PK is now (date, instrument_serial)')
        _uids = [r['uid'] for r in _cx.execute(
            'SELECT uid FROM generated_reports ORDER BY id').fetchall()]
        check(all(_uids) and len(set(_uids)) == 2,
              'existing reports are backfilled with distinct uids', str(_uids))
        check(_cx.execute("SELECT sql FROM sqlite_master "
                          "WHERE name='idx_generated_reports_uid'").fetchone() is not None,
              'and the uid unique index exists')
        _cx.close()
        _err = _migrate_only(_p15b)
        check(_err is None, 'the migration is idempotent', repr(_err))
        _cx = _sq.connect(_p15b); _cx.row_factory = _sq.Row
        check([r['uid'] for r in _cx.execute(
                  'SELECT uid FROM generated_reports ORDER BY id').fetchall()] == _uids,
              'a second migration keeps the same uids (they are identity, not state)')
        check(_cx.execute('SELECT COUNT(*) FROM weather').fetchone()[0] == 2,
              'and the weather rows are not duplicated')
        _cx.close()
        # a weather table with no date column cannot be re-keyed automatically
        _pbad15 = _build('bad_weather.db',
            'CREATE TABLE weather(session_id INTEGER PRIMARY KEY, wind_speed REAL);')
        _err = _migrate_only(_pbad15)
        check(type(_err).__name__ == 'MigrationUnsafe',
              'an unmappable weather table refuses the migration', repr(_err))

        # ── 15c. weather gap-fill on the sync-pull path ───────────────────────
        print('\n15c. the receiving Pi fills weather gaps after a sync pull')
        # A session can arrive by sync without weather (the sender had not
        # fetched). sync_peer.py runs fill_weather_gaps() after importing;
        # the fetch itself is stubbed here.
        yc = Side(os.path.join(tmp, 'wp9-pull.db'))   # constructed last: the
        # lazy noise_db import inside weather.py binds to the newest Side
        _plnc = json.loads(json.dumps(_pl15))
        for _s15 in _plnc:
            _s15.pop('wx', None)
        yc.db.import_sessions(_plnc)
        check(yc.sql('SELECT COUNT(*) FROM weather')[0][0] == 0,
              'the pull landed sessions but no weather')
        import weather as _wxmod15
        _calls15 = []
        def _fake_fetch15(date, lat, lng):
            _calls15.append((date, lat, lng))
            return {'wind_speed': 1.0, 'wind_dir': 2.0, 'temp_min': 3.0,
                    'temp_max': 4.0, 'precip': 0.0, 'hourly_json': '{"fake":true}'}
        _filled15 = _wxmod15.fill_weather_gaps(_plnc, fetch=_fake_fetch15)
        check(sorted(_filled15) == [(SESSION_DATE, 'WX-A'), (SESSION_DATE, 'WX-B')],
              'the gap-fill fetches once per weatherless session', str(_filled15))
        check(sorted(_calls15) == [(SESSION_DATE, 51.5072, -0.1276),
                                   (SESSION_DATE, 53.7997, -1.5492)],
              "each fetch uses that session's own stored coordinates", str(_calls15))
        check(yc.db.get_weather(SESSION_DATE, 'WX-B')['hourly_json'] == '{"fake":true}',
              'and stores the row under the session serial')
        _filled15 = _wxmod15.fill_weather_gaps(_plnc, fetch=_fake_fetch15)
        check(_filled15 == [] and len(_calls15) == 2,
              'a second pass refetches nothing — the rows exist now')
        # a session with no coordinates is skipped, not an error
        _sN = _payload_copy('WX-N')
        yc.db.import_sessions([_sN])
        check(_wxmod15.fill_weather_gaps([_sN], fetch=_fake_fetch15) == []
              and len(_calls15) == 2,
              'a session without coordinates is skipped')
        # a failing fetch is logged and skipped, never raised
        def _boom15(date, lat, lng):
            raise RuntimeError('api down (test)')
        yc.exec('DELETE FROM weather')
        check(_wxmod15.fill_weather_gaps(_plnc, fetch=_boom15) == [],
              'a failing fetch is contained (best-effort)')
        _sp15 = open(os.path.join(REPO, 'sync_peer.py'), encoding='utf-8').read()
        check('fill_weather_gaps(sessions)' in _sp15
              and _sp15.index('= import_sessions(sessions)')
                  < _sp15.index('fill_weather_gaps(sessions)'),
              'sync_peer.py gap-fills right after importing the pull')

        print(f'\nAll {_checks} checks passed.')
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else MEAS_DEFAULT)
