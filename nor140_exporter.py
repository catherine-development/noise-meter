"""
Build NOR140-compatible GLOBAL and PROFILE xlsx files from a run row dict.

Usage:
    from nor140_exporter import build_global_xlsx, build_profile_xlsx
    wb_bytes = build_global_xlsx(run, serial='6899108')
    wb_bytes = build_profile_xlsx(run, serial='6899108')

`run` is a dict as returned by noise_db.get_full_run_row() (includes session_date).
"""

import io
import json
from datetime import datetime, timedelta

from openpyxl import Workbook

from nor140_format import FREQ_LABELS as _FREQ_LABELS
from nor140_format import round_half_up

# 18 spectral sheets: sheet index 0–17 (Nortfr order)
_SPECTRAL_SHEETS = [
    ('LfF,99.0%_8',  'spec_lff_l99'),
    ('LfF,95.0%_7',  'spec_lff_l95'),
    ('LfF,90.0%_6',  'spec_lff_l90'),
    ('LfF,50.0%_5',  'spec_lff_l50'),
    ('LfF,10.0%_4',  'spec_lff_l10'),
    ('LfF,5.0%_3',   'spec_lff_l5'),
    ('LfF,1.0%_2',   'spec_lff_l1'),
    ('LfF,0.1%_1',   'spec_lff_l01'),
    ('LfIE',         'spec_lfie'),
    ('LfImin',       'spec_lfimin'),
    ('LfImax',       'spec_lfimax'),
    ('LfIeq',        'spec_lfieq'),
    ('LfSmin',       'spec_lfsmin'),   # capital S — matches Nortfr
    ('LfSmax',       'spec_lfsmax'),
    ('LfE',          'spec_lfe'),
    ('LfFmin',       'spec_lffmin'),   # capital F — matches Nortfr
    ('LfFmax',       'spec_lffmax'),
    ('Lfeq',         'spec_lfeq'),
]

# 16 LAF/LCF percentile scalar sheets: sheets 18–33 (descending order: 99%→0.1%)
_SCALAR_PCT_SHEETS = [
    ('LCF,99.0%_8', 'lc_l99'),
    ('LCF,95.0%_7', 'lc_l95'),
    ('LCF,90.0%_6', 'lc_l90'),
    ('LCF,50.0%_5', 'lc_l50'),
    ('LCF,10.0%_4', 'lc_l10'),
    ('LCF,5.0%_3',  'lc_l5'),
    ('LCF,1.0%_2',  'lc_l1'),
    ('LCF,0.1%_1',  'lc_l01'),
    ('LAF,99.0%_8', 'la_l99'),
    ('LAF,95.0%_7', 'la_l95'),
    ('LAF,90.0%_6', 'la_l90'),
    ('LAF,50.0%_5', 'la_l50'),
    ('LAF,10.0%_4', 'la_l10'),
    ('LAF,5.0%_3',  'la_l5'),
    ('LAF,1.0%_2',  'la_l1'),
    ('LAF,0.1%_1',  'la_l01'),
]

# 22 broadband scalar sheets: sheets 34–55
_SCALAR_BB_SHEETS = [
    ('LCpeak',  'lcpeak'),
    ('LCIE',    'lcie'),
    ('LCE',     'lce'),
    ('LCIeq',   'lcieq'),
    ('LCeq',    'lceq'),
    ('LCImin',  'lcimin'),
    ('LCSmin',  'lcsmin'),
    ('LCFmin',  'lcfmin'),
    ('LCImax',  'lcimax'),
    ('LCSmax',  'lcsmax'),
    ('LCFmax',  'lcfmax'),
    ('LApeak',  'lapeak'),
    ('LAIE',    'laie'),
    ('LAE',     'lae'),
    ('LAIeq',   'laieq'),
    ('LAeq',    'avg_laeq'),
    ('LAImin',  'laimin'),
    ('LASmin',  'lasmin'),
    ('LAFmin',  'lafmin'),
    ('LAImax',  'laimax'),
    ('LASmax',  'lasmax'),
    ('LAFmax',  'lafmax'),
]

# PROFILE channels: sheets 0–4
_PROF_SHEETS = [
    ('LApeak',  'prof_lapeak_json'),
    ('LAE',     'prof_lae_json'),
    ('LAFmax',  'prof_lafmax_json'),
    ('LAeq',    'prof_laeq_json'),
    ('LAFspl',  'prof_lafspl_json'),
]

# Five spectral groups included in the GLOBAL Summary sheet (180 extra columns)
_SUMMARY_SPECTRAL = [
    ('Lfeq',   'spec_lfeq'),
    ('LfFmax', 'spec_lffmax'),
    ('LfFmin', 'spec_lffmin'),
    ('LfE',    'spec_lfe'),
    ('LfSmax', 'spec_lfsmax'),
]

# Summary sheet scalar column order
_SUMMARY_SCALARS = [
    ('LAFmax',  'lafmax'),  ('LASmax',  'lasmax'),  ('LAImax',  'laimax'),
    ('LAFmin',  'lafmin'),  ('LASmin',  'lasmin'),  ('LAImin',  'laimin'),
    ('LAeq',    'avg_laeq'), ('LAIeq',  'laieq'),   ('LAE',     'lae'),
    ('LAIE',    'laie'),    ('LApeak',  'lapeak'),
    ('LCFmax',  'lcfmax'),  ('LCSmax',  'lcsmax'),  ('LCImax',  'lcimax'),
    ('LCFmin',  'lcfmin'),  ('LCSmin',  'lcsmin'),  ('LCImin',  'lcimin'),
    ('LCeq',    'lceq'),    ('LCIeq',   'lcieq'),   ('LCE',     'lce'),
    ('LCIE',    'lcie'),    ('LCpeak',  'lcpeak'),
    ('LAF,0.1%_1', 'la_l01'), ('LAF,1.0%_2', 'la_l1'),  ('LAF,5.0%_3',  'la_l5'),
    ('LAF,10.0%_4','la_l10'), ('LAF,50.0%_5','la_l50'), ('LAF,90.0%_6', 'la_l90'),
    ('LAF,95.0%_7','la_l95'), ('LAF,99.0%_8','la_l99'),
    ('LCF,0.1%_1', 'lc_l01'), ('LCF,1.0%_2', 'lc_l1'),  ('LCF,5.0%_3',  'lc_l5'),
    ('LCF,10.0%_4','lc_l10'), ('LCF,50.0%_5','lc_l50'), ('LCF,90.0%_6', 'lc_l90'),
    ('LCF,95.0%_7','lc_l95'), ('LCF,99.0%_8','lc_l99'),
]


# ── Datetime / duration helpers ──────────────────────────────────────────────

def _dt(run):
    """Extract start datetime from a run dict.
    start_time may be a bare time ('23:27:36') or a full ISO datetime string.
    """
    st = run.get('start_time', '') or ''
    if len(st) <= 8 and ':' in st:
        date = run.get('session_date', '1970-01-01')
        st = f'{date}T{st}'
    for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%SZ'):
        try:
            return datetime.strptime(st, fmt)
        except ValueError:
            pass
    return datetime.fromisoformat(st.replace('Z', ''))


def _run_num(run):
    """The run's identifier for filenames: the PROJ folder number from the SD card.

    NOT runs.run_number, which is a sequential index assigned at import. The two
    diverge as soon as a date has gaps or rejected runs — 2025-07-12 has no
    PROJ0020 or PROJ0023, so its PROJ0024 is run_number 22, and naming the export
    _0022 does not match what Nortfr produces.
    """
    src = (run.get('source_file') or '').strip().upper()
    if src.startswith('PROJ') and src[4:].isdigit():
        return int(src[4:])
    return run.get('run_number', 1) or 1


def _fmt_dt(dt):
    """Format datetime for DATA rows: (2026-08-12 23:27:36.000)"""
    return dt.strftime('(%Y-%m-%d %H:%M:%S.000)')


def _fmt_trig_header(dt):
    """Format datetime for the Trig time HEADER cell: (2026/8/12 23:27:36.0)"""
    # Unpadded on every component, matching Nortfr: (2026/8/12 12:8:30.0).
    # Padding the time here silently matched until a run started before 10
    # past the hour.
    return f'({dt.year}/{dt.month}/{dt.day} {dt.hour}:{dt.minute}:{dt.second}.0)'


def _fmt_duration(seconds):
    """Format duration as (H:M:S.0) — no zero-padding on any component."""
    total = int(seconds)
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    return f'({h}:{m}:{s}.0)'


_NO_DATA = '-'


def _rv(value):
    """Round a dB value to 1 decimal (half-up, matching instrument convention).

    A raw word of 0 decodes to exactly -20 dB, which the meter uses as a
    "band not measured" sentinel — it appears in spectral tables for short or
    quiet runs. Nortfr renders those cells as '-', not as -20, so emitting the
    number would both look wrong and read as a real (absurd) level.
    """
    # None now arrives here too: the sentinel is dropped at the decode boundary
    # (nor140_format.NO_DATA_DB), so an unrecorded statistic reaches the export
    # as None rather than -20. Both mean "the meter wrote nothing", and Nortfr
    # renders that as '-'. The <= test stays for rows stored before the change.
    if value is None or value <= -19.99:
        return _NO_DATA
    return round_half_up(value, 1)


def _as_list(value):
    """Coerce a spec_*/prof_* field to a list, accepting either a JSON string
    (as stored in the DB) or an already-parsed list (as returned directly by
    noise_parser, e.g. in a direct parse-to-export smoke test)."""
    if isinstance(value, str):
        return json.loads(value)
    return value or []


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_period_trigger_block(ws, n, duration_s, trig_dt, period_s=None):
    """Write the 6 standard period/trigger/duration rows shared by every sheet type:

      Period length       | period_dur | H:M:S.mS   (one period; = total for GLOBAL)
      Total periods        | n
      Before trigger        | 0
      After trigger          | n
      Trig time            | trig_hdr   | Y-Mo-D H:M:S.mS
      Effective duration    | total_dur  | H:M:S.mS   (whole run)

    period_s: duration of one period in seconds (defaults to duration_s).
              Set to 1 for PROFILE sheets where each row = 1 second.
    """
    if period_s is None:
        period_s = duration_s
    period_dur = _fmt_duration(period_s)
    total_dur = _fmt_duration(duration_s)
    trig_hdr = _fmt_trig_header(trig_dt)
    ws.append(['Period length', period_dur, 'H:M:S.mS'])
    ws.append(['Total number of periods', n])
    ws.append(['Number of periods before trigger', 0])
    ws.append(['Number of periods after trigger', n])
    ws.append(['Trig time', trig_hdr, 'Y-Mo-D H:M:S.mS'])
    ws.append(['Measurement effective duration', total_dur, 'H:M:S.mS'])


def _write_header(ws, n, duration_s, trig_dt, label=None, period_s=None):
    """Write 8-row header block (rows 0-7): the period/trigger block plus a
    blank row and a label row (or two blank rows if no label)."""
    _write_period_trigger_block(ws, n, duration_s, trig_dt, period_s=period_s)
    ws.append([])
    if label is not None:
        ws.append([None, None, None, label])
    else:
        ws.append([])


def _write_spectral_sheet(ws, label, spec_json, trig_dt, duration_s, period_s):
    """Write one spectral (36-band) sheet.

    Row 7: [None, None, None, label]
    Row 8: ['Period:', 'Time:', None, '6.3 Hz', '8.0 Hz', ...]
    Row 9: [0, datetime, None, val0, val1, ...]
    """
    _write_header(ws, 1, duration_s, trig_dt, label=label, period_s=period_s)
    ws.append(['Period:', 'Time:', None] + _FREQ_LABELS)
    vals = _as_list(spec_json)
    ws.append([0, _fmt_dt(trig_dt), None] + [_rv(v) for v in vals])


def _write_scalar_sheet(ws, label, value, trig_dt, duration_s, period_s):
    """Write one scalar sheet.

    Row 7: (blank)
    Row 8: ['Period:', 'Time:', None, label]
    Row 9: [0, datetime, None, value]
    """
    _write_header(ws, 1, duration_s, trig_dt, label=None, period_s=period_s)
    ws.append(['Period:', 'Time:', None, label])
    ws.append([0, _fmt_dt(trig_dt), None, _rv(value)])


def _write_prof_sheet(ws, label, prof_json, trig_dt, duration_s):
    """Write one PROFILE 1-second time-series sheet.

    PROF data is always stored at 1-second resolution, so period = 1 s.
    Row 7: (blank)
    Row 8: ['Period:', 'Time:', None, label]
    Row 9+: [i, datetime_i, None, value_i]
    """
    vals = _as_list(prof_json)
    n = len(vals)
    _write_header(ws, n, duration_s, trig_dt, label=None, period_s=1)  # one period = 1 s
    ws.append(['Period:', 'Time:', None, label])
    for i, v in enumerate(vals):
        t = trig_dt + timedelta(seconds=i)
        ws.append([i, _fmt_dt(t), None, _rv(v)])


def _write_global_summary(ws, run, trig_dt, duration_s):
    """Write GLOBAL Summary sheet — verified cell-for-cell against the Nortfr
    reference workbook (NOR140_6899108_260812_0009_GLOBAL.xlsx):

      Row 1: ['Period:', 'Time:', 'Duration:', None, None, <38 scalar headers>,
              <group label in the FIRST of its 36 cols, None for the other 35>] x5 groups
      Row 2: blank
      Row 3: blank under Period/Time/Duration/scalars, then <36 freq labels> x5 groups
      Row 4: [0, datetime, None, None, None, <38 scalar vals>, <36 band vals> x5 groups]
      Row 5-6: blank
      Row 7-9: NC / NR / RC II room-acoustics ratings

    Column count: 5 + 38 + 180 = 223, matching Nortfr.

    NC/NR/RC ratings (rows 7-9) are room-acoustics curve fits Nortfr computes
    from the spectrum; this app doesn't implement that curve-fitting, and the
    reference workbook itself shows '-' for all three on this (outdoor,
    non-room) measurement, so that's what's written here. A genuinely
    room-acoustics measurement would need real NC/NR/RC computation to match.
    """
    scalar_headers = [name for name, _ in _SUMMARY_SCALARS]
    row1 = ['Period:', 'Time:', 'Duration:', None, None] + scalar_headers
    for group, _ in _SUMMARY_SPECTRAL:
        row1.append(group)
        row1.extend([None] * (len(_FREQ_LABELS) - 1))
    ws.append(row1)

    ws.append([])

    row3 = [None] * (5 + len(_SUMMARY_SCALARS))
    for _ in _SUMMARY_SPECTRAL:
        row3.extend(_FREQ_LABELS)
    ws.append(row3)

    scalar_vals = [_rv(run.get(col)) for _, col in _SUMMARY_SCALARS]
    spec_vals = []
    for _, col in _SUMMARY_SPECTRAL:
        raw = run.get(col)
        bands = json.loads(raw) if isinstance(raw, str) else (raw or [None] * len(_FREQ_LABELS))
        spec_vals.extend(_rv(v) for v in bands)
    ws.append([0, _fmt_dt(trig_dt), None, None, None] + scalar_vals + spec_vals)

    ws.append([])
    ws.append([])
    ws.append(['NC', '-', 'dB'])
    ws.append(['NR', '-', 'dB'])
    ws.append(['RC II (-)', '-', 'dB'])


def _write_setup(ws, run, serial, report_type, trig_dt, n, duration_s, period_s=None):
    """Write a Setup sheet (GLOBAL or PROFILE), matching Nortfr's 14-row layout:

      (blank)
      File version
      Filename            e.g. NOR140_<serial>_<date>_<run>.NBF (NOR140_<serial>_<date>_<run>_<TYPE>)
      Report type
      Bandwidth
      Frequency range
      <6-row period/trigger block — same structure as every other sheet>
      Full scale
      Sensitivity
    """
    date_str = trig_dt.strftime('%y%m%d')
    run_num = _run_num(run)
    base = f'NOR140_{serial}_{date_str}_{run_num:04d}'
    ws.append([])
    ws.append(['File version', 'v1.0/6.1.1.51'])
    ws.append(['Filename', f'{base}.NBF ({base}_{report_type})'])
    ws.append(['Report type', report_type])
    ws.append(['Bandwidth', '1/3 Octave'])
    ws.append(['Frequency range', '6.3 Hz - 20.0 kHz'])
    _write_period_trigger_block(ws, n, duration_s, trig_dt, period_s=period_s)
    # Full scale is a per-measurement range setting (90 / 100 / 130 dB seen in
    # this archive), read from the GLOB. Sensitivity also varies (-26.9, -27.0,
    # -27.4 seen) but its offset has not been located, so it stays a default and
    # will differ from Nortfr on instruments calibrated to another value.
    ws.append(['Full scale', run.get('full_scale') or 130, 'dB'])
    ws.append(['Sensitivity', run.get('sensitivity_db') or -26.9, 'dB'])


def _write_global_setup(ws, run, serial, trig_dt, duration_s, period_s):
    """Write GLOBAL Setup sheet (one period covering the whole run).

    period_s is the record count, duration_s the meter's stored duration; they
    differ when a run was stopped mid-period.
    """
    _write_setup(ws, run, serial, 'GLOBAL', trig_dt, n=1, duration_s=duration_s,
                 period_s=period_s)


def _profile_n(run):
    prof_json = run.get('prof_laeq_json') or run.get('prof_lafspl_json')
    return len(_as_list(prof_json)) if prof_json else (run.get('n_samples', 1) or 1)


def _write_profile_summary(ws, run, trig_dt):
    """Write PROFILE Summary sheet: the full 1-second time series of all 5
    PROF channels side by side, in reverse sheet order (LAFspl, LAeq, LAFmax,
    LAE, LApeak), followed by two blank rows and three NC/NR/RC room-acoustics
    rating rows — verified against the Nortfr reference (908 rows total).
    """
    channels = list(reversed(_PROF_SHEETS))
    headers = [name for name, _ in channels]
    ws.append(['Period:', 'Time:', 'Duration:', None, None] + headers)
    ws.append([])
    ws.append([])
    series = []
    n = 0
    for _, col in channels:
        vals = _as_list(run.get(col))
        series.append(vals)
        n = max(n, len(vals))
    for i in range(n):
        t = trig_dt + timedelta(seconds=i)
        row_vals = [_rv(s[i]) if i < len(s) else None for s in series]
        ws.append([i, _fmt_dt(t), None, None, None] + row_vals)

    ws.append([])
    ws.append([])
    ws.append(['NC', '-', 'dB'])
    ws.append(['NR', '-', 'dB'])
    ws.append(['RC II (-)', '-', 'dB'])


def _write_profile_setup(ws, run, serial, trig_dt, duration_s):
    """Write PROFILE Setup sheet (one period per second, for the run's full duration)."""
    n = _profile_n(run)
    _write_setup(ws, run, serial, 'PROFILE', trig_dt, n=n, duration_s=duration_s, period_s=1)


# ── Public API ────────────────────────────────────────────────────────────────

def build_global_xlsx(run, serial=''):
    """Build NOR140-style GLOBAL xlsx. Returns bytes."""
    trig_dt = _dt(run)
    n_samples = run.get('n_samples', 1) or 1
    # n_samples is the count of actual 1-second measurements; 'step' is a chart
    # downsampling factor and must NOT multiply the real measurement duration.
    # Prefer the duration the meter stored: a run stopped mid-period writes a
    # final partial record, so the record count can exceed the elapsed time
    # (run 1 of 2026-08-12 has 83 records but ran 0:1:22).
    duration_s = run.get('duration_s') or n_samples

    wb = Workbook()
    wb.remove(wb.active)

    for name, col in _SPECTRAL_SHEETS:
        ws = wb.create_sheet(name)
        _write_spectral_sheet(ws, name, run.get(col) or '[]', trig_dt, duration_s, n_samples)

    for name, col in _SCALAR_PCT_SHEETS:
        ws = wb.create_sheet(name)
        _write_scalar_sheet(ws, name, run.get(col), trig_dt, duration_s, n_samples)

    for name, col in _SCALAR_BB_SHEETS:
        ws = wb.create_sheet(name)
        _write_scalar_sheet(ws, name, run.get(col), trig_dt, duration_s, n_samples)

    ws = wb.create_sheet('Summary')
    _write_global_summary(ws, run, trig_dt, duration_s)

    ws = wb.create_sheet('Setup')
    _write_global_setup(ws, run, serial, trig_dt, duration_s, n_samples)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_profile_xlsx(run, serial=''):
    """Build NOR140-style PROFILE xlsx. Returns bytes."""
    trig_dt = _dt(run)
    # Prefer the meter's stored duration over the record count — see the note in
    # build_global_xlsx.
    duration_s = run.get('duration_s') or (run.get('n_samples', 1) or 1)

    wb = Workbook()
    wb.remove(wb.active)

    for name, col in _PROF_SHEETS:
        ws = wb.create_sheet(name)
        _write_prof_sheet(ws, name, run.get(col) or '[]', trig_dt, duration_s)

    ws = wb.create_sheet('Summary')
    _write_profile_summary(ws, run, trig_dt)

    ws = wb.create_sheet('Setup')
    _write_profile_setup(ws, run, serial, trig_dt, duration_s)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(run, serial, report_type):
    """Return the canonical Nortfr filename for a run export."""
    trig_dt = _dt(run)
    date_str = trig_dt.strftime('%y%m%d')
    run_num = _run_num(run)
    return f'NOR140_{serial}_{date_str}_{run_num:04d}_{report_type}.xlsx'
